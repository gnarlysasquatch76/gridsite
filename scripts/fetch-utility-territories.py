"""
Fetch HIFLD Electric Retail Service Territory polygons, join with EIA 860
generation capacity and EIA 861 sales data, compute generation-to-load
ratios, and output a color-coded GeoJSON for the map layer.

Sources:
  - HIFLD Electric Retail Service Territories (ArcGIS FeatureServer)
  - EIA Form 860 (december_generator2025.xlsx) — generation capacity
  - EIA Form 861 (f8612024.zip) — annual sales to ultimate customers
"""

import json
import os
import re
import io
import urllib.request
import urllib.parse
import zipfile
import time
import openpyxl

SCRIPT_DIR = os.path.dirname(__file__)
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "..", "public", "data", "utility-territories.geojson")
EIA_860_FILE = os.path.join(SCRIPT_DIR, "..", "data", "december_generator2025.xlsx")
EIA_861_URL = "https://www.eia.gov/electricity/data/eia861/zip/f8612024.zip"

HIFLD_URL = "https://services3.arcgis.com/OYP7N6mAJJCyH6hd/arcgis/rest/services/Electric_Retail_Service_Territories_HIFLD/FeatureServer/0/query"
PAGE_SIZE = 2000

US_STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA",
    "HI","ID","IL","IN","IA","KS","KY","LA","ME","MD",
    "MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC",
    "SD","TN","TX","UT","VT","VA","WA","WV","WI","WY",
    "DC","PR","VI","GU","AS","MP",
}


def safe_float(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def normalize_name(name):
    """Normalize utility name for fuzzy matching."""
    if not name:
        return ""
    s = name.upper().strip()
    # Remove common suffixes
    for suffix in [" INC", " LLC", " CO", " CORP", " CORPORATION",
                   " COMPANY", " LTD", " LP", " L.P.", " L.L.C.",
                   " INCORPORATED", " AUTHORITY", " AUTH"]:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
    # Remove punctuation
    s = re.sub(r"[.,'\"-]", "", s)
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ── Step 1: Fetch HIFLD Territory Polygons ──────────────────────────────

def fetch_page(offset):
    """Fetch a single page of territory polygons via POST."""
    params = urllib.parse.urlencode({
        "where": "1=1",
        "outFields": "NAME,STATE,TYPE,CUSTOMERS,NAICS_DESC,SUMMER_CAP,RETAIL_MWH,NET_GEN",
        "outSR": "4326",
        "f": "geojson",
        "resultRecordCount": str(PAGE_SIZE),
        "resultOffset": str(offset),
        "maxAllowableOffset": "0.005",
    }).encode("utf-8")

    for attempt in range(3):
        try:
            req = urllib.request.Request(HIFLD_URL, data=params, headers={"User-Agent": "GridSite/1.0"})
            with urllib.request.urlopen(req, timeout=180) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            if "error" in data:
                raise Exception("API error: " + str(data["error"]))
            return data
        except Exception as e:
            print("  Attempt " + str(attempt + 1) + " failed: " + str(e))
            if attempt < 2:
                time.sleep(3 * (attempt + 1))
            else:
                raise


def fetch_territories():
    """Fetch all HIFLD territory polygons with pagination."""
    all_features = []
    offset = 0

    print("Step 1: Fetching HIFLD Electric Retail Service Territories...")
    print("  Source: " + HIFLD_URL.split("/query")[0])
    print()

    while True:
        print("  Fetching offset " + str(offset) + "...")
        data = fetch_page(offset)

        features = data.get("features", [])
        if len(features) == 0:
            break

        # Filter to US states
        for f in features:
            props = f.get("properties", {})
            state = (props.get("STATE") or "").strip().upper()
            if state in US_STATES:
                all_features.append(f)

        print("    Got " + str(len(features)) + " records, total kept " + str(len(all_features)))

        if len(features) < PAGE_SIZE:
            break

        offset += PAGE_SIZE
        time.sleep(0.5)

    print("  Total territories: " + str(len(all_features)))
    return all_features


# ── Step 2: Read EIA 860 Generator Data ─────────────────────────────────

def read_eia_860():
    """Read EIA 860 and aggregate nameplate MW by utility/entity name."""
    print()
    print("Step 2: Reading EIA 860 generator data...")
    print("  File: " + EIA_860_FILE)

    wb = openpyxl.load_workbook(EIA_860_FILE, read_only=True)
    ws = wb["Operating"]

    # Column 0 = Entity ID, Column 1 = Entity Name, Column 12 = Nameplate MW
    capacity_by_name = {}
    header_rows = 3
    row_count = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx < header_rows:
            continue

        entity_name = str(row[1] or "").strip()
        mw = safe_float(row[12])

        if not entity_name or mw is None:
            continue

        row_count += 1
        key = normalize_name(entity_name)
        capacity_by_name[key] = capacity_by_name.get(key, 0.0) + mw

    wb.close()

    print("  Generator rows processed: " + str(row_count))
    print("  Unique utilities with capacity: " + str(len(capacity_by_name)))
    return capacity_by_name


# ── Step 3: Download & Read EIA 861 Sales Data ──────────────────────────

def read_eia_861():
    """Download EIA 861 zip and aggregate annual MWh sales by utility name."""
    print()
    print("Step 3: Downloading EIA 861 sales data...")
    print("  URL: " + EIA_861_URL)

    req = urllib.request.Request(EIA_861_URL, headers={"User-Agent": "GridSite/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        zip_data = resp.read()

    print("  Downloaded " + str(round(len(zip_data) / 1024 / 1024, 1)) + " MB")

    zf = zipfile.ZipFile(io.BytesIO(zip_data))
    sales_file = None
    for name in zf.namelist():
        if "sales" in name.lower() and name.lower().endswith(".xlsx"):
            sales_file = name
            break

    if not sales_file:
        # Fallback: try csv
        for name in zf.namelist():
            if "sales" in name.lower() and name.lower().endswith(".csv"):
                sales_file = name
                break

    if not sales_file:
        print("  WARNING: Could not find sales file in ZIP. Files: " + str(zf.namelist()))
        return {}

    print("  Reading: " + sales_file)

    sales_by_name = {}

    if sales_file.lower().endswith(".xlsx"):
        xlsx_data = zf.read(sales_file)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)
        ws = wb.active

        # Find header row and column indices
        header_row_idx = None
        col_utility_name = None
        col_ownership = None
        mwh_cols = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_vals = [str(v or "").strip().upper() for v in row]

            # Look for the header row containing "UTILITY NAME" or "ENTITY NAME"
            for ci, val in enumerate(row_vals):
                if "UTILITY" in val and "NAME" in val:
                    header_row_idx = row_idx
                    col_utility_name = ci
                    break
                if "ENTITY" in val and "NAME" in val:
                    header_row_idx = row_idx
                    col_utility_name = ci
                    break

            if header_row_idx is not None:
                # Find ownership/type column and MWh columns
                for ci, val in enumerate(row_vals):
                    if "OWNERSHIP" in val or "OWNER" in val:
                        col_ownership = ci
                    # Look for total sales MWh columns (TOTAL column with MEGAWATTHOURS)
                    if "MEGAWATTHOUR" in val and "TOTAL" in val:
                        mwh_cols.append(ci)
                    elif "TOTAL" in val and "MWH" in val:
                        mwh_cols.append(ci)

                # If no specific total MWh column, look for any MEGAWATTHOURS columns
                if not mwh_cols:
                    for ci, val in enumerate(row_vals):
                        if "MEGAWATTHOUR" in val or ("MWH" in val and "REVENUE" not in val):
                            mwh_cols.append(ci)

                break

        if col_utility_name is None:
            print("  WARNING: Could not find utility name column in 861 sales data")
            wb.close()
            return {}

        print("  Utility name column: " + str(col_utility_name))
        print("  Ownership column: " + str(col_ownership))
        print("  MWh columns: " + str(mwh_cols))

        row_count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx <= header_row_idx:
                continue

            if len(row) <= col_utility_name:
                continue

            entity_name = str(row[col_utility_name] or "").strip()
            if not entity_name:
                continue

            # Sum all MWh columns for total sales
            total_mwh = 0.0
            for ci in mwh_cols:
                if ci < len(row):
                    val = safe_float(row[ci])
                    if val is not None and val > 0:
                        total_mwh += val

            ownership = ""
            if col_ownership is not None and col_ownership < len(row):
                ownership = str(row[col_ownership] or "").strip()

            key = normalize_name(entity_name)
            if key not in sales_by_name:
                sales_by_name[key] = {"total_mwh": 0.0, "utility_type": ownership}
            sales_by_name[key]["total_mwh"] += total_mwh
            row_count += 1

        wb.close()
        print("  Sales rows processed: " + str(row_count))

    zf.close()
    print("  Unique utilities with sales data: " + str(len(sales_by_name)))
    return sales_by_name


# ── Step 4: Join & Compute Ratios ────────────────────────────────────────

def join_and_compute(territories, capacity_by_name, sales_by_name):
    """Join HIFLD territories with EIA data and compute ratios.

    Uses HIFLD embedded capacity/sales as primary source, then overrides
    with EIA 860/861 data where available for better accuracy.
    """
    print()
    print("Step 4: Joining data and computing ratios...")

    matched_cap = 0
    matched_sales = 0
    used_hifld_cap = 0
    used_hifld_sales = 0
    ratio_counts = {"surplus": 0, "balanced": 0, "constrained": 0, "unknown": 0}

    for feature in territories:
        props = feature.get("properties", {})
        name = (props.get("NAME") or "Unknown").strip()
        state = (props.get("STATE") or "").strip()
        customers = safe_float(props.get("CUSTOMERS"))
        naics = (props.get("NAICS_DESC") or "").strip()
        util_type = (props.get("TYPE") or "").strip()

        # HIFLD embedded data
        hifld_cap = safe_float(props.get("SUMMER_CAP"))
        hifld_retail_mwh = safe_float(props.get("RETAIL_MWH"))

        key = normalize_name(name)

        # EIA 860 capacity (preferred over HIFLD)
        eia_cap = capacity_by_name.get(key)

        # EIA 861 sales (preferred over HIFLD)
        sales_info = sales_by_name.get(key, {})
        eia_mwh = sales_info.get("total_mwh")
        eia_type = sales_info.get("utility_type", "")

        # Use EIA data if available, fall back to HIFLD
        capacity_mw = eia_cap
        if capacity_mw is not None:
            matched_cap += 1
        elif hifld_cap and hifld_cap > 0:
            capacity_mw = hifld_cap
            used_hifld_cap += 1

        total_mwh = eia_mwh if eia_mwh and eia_mwh > 0 else None
        if total_mwh:
            matched_sales += 1
        elif hifld_retail_mwh and hifld_retail_mwh > 0:
            total_mwh = hifld_retail_mwh
            used_hifld_sales += 1

        utility_type = eia_type or util_type or naics

        # Compute ratio
        avg_load_mw = None
        ratio = None
        ratio_class = "unknown"

        if capacity_mw is not None and capacity_mw > 0 and total_mwh and total_mwh > 0:
            avg_load_mw = total_mwh / 8760.0
            ratio = capacity_mw / avg_load_mw

            if ratio > 1.5:
                ratio_class = "surplus"
            elif ratio >= 0.7:
                ratio_class = "balanced"
            else:
                ratio_class = "constrained"

        ratio_counts[ratio_class] += 1

        # Replace properties with clean output
        feature["properties"] = {
            "name": name,
            "state": state,
            "utility_type": utility_type,
            "customers": int(customers) if customers else None,
            "capacity_mw": round(capacity_mw, 1) if capacity_mw is not None else None,
            "sales_mwh": round(total_mwh, 0) if total_mwh else None,
            "avg_load_mw": round(avg_load_mw, 1) if avg_load_mw is not None else None,
            "ratio": round(ratio, 2) if ratio is not None else None,
            "ratio_class": ratio_class,
        }

    print("  Territories matched to EIA 860 capacity: " + str(matched_cap) + " / " + str(len(territories)))
    print("  Territories using HIFLD capacity: " + str(used_hifld_cap))
    print("  Territories matched to EIA 861 sales: " + str(matched_sales) + " / " + str(len(territories)))
    print("  Territories using HIFLD retail MWh: " + str(used_hifld_sales))
    print("  Ratio breakdown:")
    for cls in ["surplus", "balanced", "constrained", "unknown"]:
        print("    " + cls + ": " + str(ratio_counts[cls]))


# ── Step 5: Output GeoJSON ──────────────────────────────────────────────

def main():
    # Step 1: Fetch territories
    territories = fetch_territories()

    # Step 2: Read EIA 860
    capacity_by_name = read_eia_860()

    # Step 3: Download & read EIA 861
    sales_by_name = read_eia_861()

    # Step 4: Join & compute
    join_and_compute(territories, capacity_by_name, sales_by_name)

    # Step 5: Write output
    print()
    print("Step 5: Writing output...")

    geojson = {
        "type": "FeatureCollection",
        "features": territories,
    }

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w") as f:
        json.dump(geojson, f)

    file_size = round(os.path.getsize(OUTPUT_FILE) / 1024 / 1024, 1)

    # Stats
    states = {}
    for feat in territories:
        st = feat.get("properties", {}).get("state", "??")
        states[st] = states.get(st, 0) + 1

    print()
    print("Done!")
    print("  Total territories: " + str(len(territories)))
    print("  States covered: " + str(len(states)))
    print("  Top 10 states: " + ", ".join(
        s + "=" + str(c) for s, c in sorted(states.items(), key=lambda x: -x[1])[:10]
    ))
    print("  Output: " + OUTPUT_FILE + " (" + str(file_size) + " MB)")


if __name__ == "__main__":
    main()
