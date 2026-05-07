"""
Stranded Capacity Pipeline — identify sites where large power consumers are
closing, leaving behind utility infrastructure sized for 30MW+ loads.

Three data ingestion strategies:
  1. WARN Act Scraper — state labor dept mass layoff/closure filings
  2. EIA Large Consumer Analysis — research agent for utility territory load loss
  3. News & SEC Filing Scanner — web search for recent industrial closures

All strategies output GeoJSON features compatible with the existing scoring model.

Usage:
    python3 scripts/stranded_capacity.py                  # Run all strategies
    python3 scripts/stranded_capacity.py --warn-only      # WARN Act only
    python3 scripts/stranded_capacity.py --news-only      # News scanner only
    python3 scripts/stranded_capacity.py --eia-only       # EIA analysis only
    python3 scripts/stranded_capacity.py --states OH,PA   # Target specific states
    python3 scripts/stranded_capacity.py --dry-run        # Show plan without executing
"""

import argparse
import asyncio
import csv
import io
import json
import math
import os
import re
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

CUTOFF_MONTHS = 24  # Only include closures from the last 24 months

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "public", "data")
WARN_OUTPUT = os.path.join(DATA_DIR, "warn-closures.geojson")
NEWS_OUTPUT = os.path.join(DATA_DIR, "industrial-closures.geojson")
EIA_OUTPUT = os.path.join(DATA_DIR, "stranded-capacity-research.json")
SUBSTATIONS_FILE = os.path.join(DATA_DIR, "substations.geojson")
TRANSMISSION_FILE = os.path.join(DATA_DIR, "transmission-lines.geojson")
DB_FILE = os.path.join(SCRIPT_DIR, "..", "stranded_capacity.db")

TARGET_STATES = ["VA", "TX", "OH", "GA", "IL", "IN", "PA", "NC", "AZ", "NV",
                 "IA", "OR", "WA", "TN", "SC", "NJ", "AL"]

INDUSTRY_KEYWORDS = [
    "manufactur", "steel", "aluminum", "smelting", "chemical", "paper",
    "automotive", "assembly", "foundry", "refinery", "glass", "cement",
    "mining", "distribution center", "fulfillment", "data center",
    "plant clos", "mill", "processing", "semiconductor", "pharma",
    "plastics", "metals", "fabricat", "warehouse", "logistics",
]

MIN_EMPLOYEES = 200
MIN_MW = 30

# Estimated MW by facility sub-type
MW_ESTIMATES = {
    "steel mill": 125, "smelter": 150, "aluminum": 175,
    "auto assembly": 35, "automotive": 35,
    "paper mill": 55, "paper": 55,
    "chemical plant": 60, "chemical": 60,
    "distribution center": 20, "fulfillment": 25, "warehouse": 15,
    "data center": 50, "refinery": 90,
    "glass": 45, "cement": 50, "foundry": 40,
    "semiconductor": 60, "pharma": 35, "manufacturing": 30,
    "processing": 30, "plastics": 25, "metals": 40, "fabricat": 30,
    "mining": 40, "logistics": 15, "mill": 45,
}

MODEL = "claude-sonnet-4-20250514"
MAX_TOKENS = 2000
CONCURRENCY = 3

INPUT_COST_PER_TOKEN = 3.0 / 1_000_000
OUTPUT_COST_PER_TOKEN = 15.0 / 1_000_000

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_DELAY = 1.1  # seconds between requests


# ── Utilities ─────────────────────────────────────────────────────────────


def parse_date(date_str):
    """Try to parse a date string into a datetime. Returns None on failure."""
    if not date_str or not date_str.strip():
        return None
    date_str = date_str.strip()
    for fmt in [
        "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
        "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
        "%B %d, %Y", "%b %d, %Y", "%d-%b-%Y",
    ]:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def is_within_cutoff(date_str):
    """Check if a date string is within the last CUTOFF_MONTHS months.
    Returns True if date is recent enough or if date can't be parsed (keep unknowns)."""
    dt = parse_date(date_str)
    if dt is None:
        return True  # keep records with unparseable dates — better to include than miss
    cutoff = datetime.now() - timedelta(days=CUTOFF_MONTHS * 30)
    return dt >= cutoff


def normalize_date(date_str):
    """Normalize a date string to YYYY-MM-DD format."""
    dt = parse_date(date_str)
    if dt is None:
        return date_str.strip() if date_str else ""
    return dt.strftime("%Y-%m-%d")


def haversine_miles(lat1, lon1, lat2, lon2):
    R = 3958.8
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def estimate_mw(text):
    """Estimate MW from facility description text."""
    text_lower = text.lower()
    for keyword, mw in sorted(MW_ESTIMATES.items(), key=lambda x: -x[1]):
        if keyword in text_lower:
            return mw
    return 30  # default for large industrial


def classify_sub_type(text):
    """Classify facility sub-type from description."""
    text_lower = text.lower()
    mappings = [
        ("steel", "Steel Mill"), ("smelter", "Smelter"), ("aluminum", "Aluminum Smelter"),
        ("auto", "Auto Assembly"), ("paper", "Paper Mill"), ("chemical", "Chemical Plant"),
        ("refiner", "Refinery"), ("glass", "Glass Plant"), ("cement", "Cement Plant"),
        ("foundry", "Foundry"), ("semiconductor", "Semiconductor Fab"),
        ("pharma", "Pharmaceutical"), ("data center", "Data Center"),
        ("distribution", "Distribution Center"), ("fulfillment", "Fulfillment Center"),
        ("warehouse", "Warehouse/Logistics"), ("logistics", "Warehouse/Logistics"),
        ("mining", "Mining Operation"), ("mill", "Industrial Mill"),
        ("processing", "Processing Plant"), ("manufactur", "Manufacturing"),
    ]
    for keyword, label in mappings:
        if keyword in text_lower:
            return label
    return "Industrial Facility"


def matches_industry(text):
    """Check if text matches any industry keyword."""
    text_lower = text.lower()
    return any(kw in text_lower for kw in INDUSTRY_KEYWORDS)


def geocode(address, state=""):
    """Geocode an address using Nominatim. Returns (lat, lon) or None."""
    query = address
    if state and state not in address:
        query = address + ", " + state
    query = query + ", USA"

    params = urllib.parse.urlencode({
        "q": query, "format": "json", "limit": "1", "countrycodes": "us",
    })
    url = NOMINATIM_URL + "?" + params
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "GridSite-StrandedCapacity/1.0 (brian@gridsite.dev)"
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print("      Geocode failed for '{}': {}".format(address, e))
    return None


# ── Database ──────────────────────────────────────────────────────────────


def init_db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS scrape_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            strategy TEXT,
            state TEXT,
            source_url TEXT,
            records_found INTEGER,
            records_matched INTEGER,
            timestamp TEXT,
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS research_log (
            site_key TEXT PRIMARY KEY,
            strategy TEXT,
            site_name TEXT,
            state TEXT,
            response_json TEXT,
            input_tokens INTEGER,
            output_tokens INTEGER,
            cost_usd REAL,
            researched_at TEXT,
            model TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS geocode_log (
            address TEXT PRIMARY KEY,
            lat REAL,
            lon REAL,
            timestamp TEXT
        )
    """)
    conn.commit()
    conn.close()


def log_scrape(strategy, state, url, found, matched, notes=""):
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO scrape_log (strategy, state, source_url, records_found, records_matched, timestamp, notes) VALUES (?,?,?,?,?,?,?)",
        (strategy, state, url, found, matched, datetime.now(timezone.utc).isoformat(), notes),
    )
    conn.commit()
    conn.close()


def get_cached_geocode(address):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.execute("SELECT lat, lon FROM geocode_log WHERE address = ?", (address,))
    row = cur.fetchone()
    conn.close()
    if row:
        return row[0], row[1]
    return None


def cache_geocode(address, lat, lon):
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT OR REPLACE INTO geocode_log (address, lat, lon, timestamp) VALUES (?,?,?,?)",
        (address, lat, lon, datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    conn.close()


def geocode_cached(address, state=""):
    """Geocode with cache layer."""
    cached = get_cached_geocode(address)
    if cached:
        return cached
    time.sleep(NOMINATIM_DELAY)
    result = geocode(address, state)
    if result:
        cache_geocode(address, result[0], result[1])
    return result


def lookup_facility_address(company, city, state):
    """Use Anthropic API to find the actual street address of a facility.
    Returns (address_string, is_approximate) or (None, True) on failure."""
    if not os.environ.get("ANTHROPIC_API_KEY"):
        return None, True

    import anthropic

    cache_key = "addr|{}|{}|{}".format(company, city, state)
    conn = sqlite3.connect(DB_FILE)
    cur = conn.execute(
        "SELECT response_json FROM research_log WHERE site_key = ?", (cache_key,))
    cached = cur.fetchone()
    if cached:
        result = json.loads(cached[0])
        conn.close()
        addr = result.get("address")
        return addr, result.get("approximate", addr is None)
    conn.close()

    prompt = (
        "Find the street address of this industrial facility. Return ONLY a JSON object.\n\n"
        "Company: {}\n"
        "City: {}, {}\n\n"
        "Return ONLY this JSON:\n"
        '{{\n'
        '  "address": "full street address including city and state" or null if not found,\n'
        '  "approximate": false if you found the exact facility address, true if guessing\n'
        '}}'
    ).format(company, city, state)

    try:
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=MODEL, max_tokens=300,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}],
        )

        text = ""
        for block in response.content:
            if hasattr(block, "text"):
                text += block.text

        input_t = response.usage.input_tokens
        output_t = response.usage.output_tokens
        cost = input_t * INPUT_COST_PER_TOKEN + output_t * OUTPUT_COST_PER_TOKEN

        result = None
        start = text.find("{")
        end = text.rfind("}") + 1
        if start >= 0 and end > start:
            try:
                result = json.loads(text[start:end])
            except json.JSONDecodeError:
                pass

        if result is None:
            result = {"address": None, "approximate": True}

        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "INSERT OR REPLACE INTO research_log (site_key, strategy, site_name, state, response_json, input_tokens, output_tokens, cost_usd, researched_at, model) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (cache_key, "address_lookup", company, state, json.dumps(result),
             input_t, output_t, cost, datetime.now(timezone.utc).isoformat(), MODEL),
        )
        conn.commit()
        conn.close()

        addr = result.get("address")
        approx = result.get("approximate", addr is None)
        return addr, approx

    except Exception as e:
        print("      Address lookup failed for {}: {}".format(company, e))
        return None, True


# ── Infrastructure proximity ─────────────────────────────────────────────


_substations = None
_transmission = None


def load_substations():
    global _substations
    if _substations is not None:
        return _substations
    if not os.path.exists(SUBSTATIONS_FILE):
        _substations = []
        return _substations
    with open(SUBSTATIONS_FILE) as f:
        geo = json.load(f)
    _substations = []
    for feat in geo["features"]:
        p = feat["properties"]
        v = p.get("MAX_VOLT")
        if v is not None and float(v) >= 138:
            coords = feat["geometry"]["coordinates"]
            _substations.append({
                "lat": float(p.get("LATITUDE", coords[1])),
                "lon": float(p.get("LONGITUDE", coords[0])),
                "max_volt": float(v),
                "name": p.get("NAME", ""),
            })
    return _substations


def find_nearest_substation(lat, lon):
    """Find nearest substation >= 138kV. Returns (name, distance_mi, voltage) or None."""
    subs = load_substations()
    best_dist = float("inf")
    best = None
    for s in subs:
        d = haversine_miles(lat, lon, s["lat"], s["lon"])
        if d < best_dist:
            best_dist = d
            best = s
    if best is None:
        return None
    return best["name"], round(best_dist, 1), best["max_volt"]


def find_nearest_hv_transmission(lat, lon):
    """Find nearest transmission line >= 345kV within 20 miles. Returns voltage or None."""
    global _transmission
    if _transmission is None:
        if not os.path.exists(TRANSMISSION_FILE):
            _transmission = []
        else:
            with open(TRANSMISSION_FILE) as f:
                geo = json.load(f)
            _transmission = []
            for feat in geo["features"]:
                v = feat["properties"].get("VOLTAGE")
                if v is not None and float(v) >= 345:
                    geom = feat["geometry"]
                    coords = []
                    if geom["type"] == "LineString":
                        coords = geom["coordinates"]
                    elif geom["type"] == "MultiLineString":
                        for seg in geom["coordinates"]:
                            coords.extend(seg)
                    if coords:
                        # Store just first and last point for rough proximity
                        _transmission.append({
                            "voltage": float(v),
                            "points": [(coords[0][1], coords[0][0]),
                                       (coords[-1][1], coords[-1][0])],
                        })

    deg_delta = 20 / 69.0
    best_voltage = None
    for t in _transmission:
        for p in t["points"]:
            if abs(p[0] - lat) > deg_delta or abs(p[1] - lon) > deg_delta:
                continue
            d = haversine_miles(lat, lon, p[0], p[1])
            if d <= 10:
                if best_voltage is None or t["voltage"] > best_voltage:
                    best_voltage = t["voltage"]
    return best_voltage


def build_feature(name, lat, lon, state, source, sub_type, company, location,
                  estimated_mw, employee_count, closure_date, closure_status,
                  sources, notes="", location_approximate=False):
    """Build a GeoJSON feature for a stranded capacity site."""
    # Infrastructure proximity
    sub_info = find_nearest_substation(lat, lon)
    nearest_sub_name = sub_info[0] if sub_info else ""
    nearest_sub_km = round(sub_info[1] * 1.60934, 1) if sub_info else 0
    nearest_sub_kv = sub_info[2] if sub_info else 0
    nearest_tx_kv = find_nearest_hv_transmission(lat, lon)

    return {
        "type": "Feature",
        "geometry": {"type": "Point", "coordinates": [lon, lat]},
        "properties": {
            "name": name,
            "source": source,
            "site_type": "Stranded Capacity",
            "sub_type": sub_type,
            "company": company,
            "location": location,
            "state": state,
            "estimated_mw": estimated_mw,
            "employee_count": employee_count,
            "closure_date": closure_date,
            "closure_status": closure_status,
            "utility": "",
            "location_approximate": location_approximate,
            "nearest_substation_name": nearest_sub_name,
            "nearest_substation_km": nearest_sub_km,
            "nearest_substation_kv": nearest_sub_kv,
            "nearest_transmission_kv": nearest_tx_kv or 0,
            "data_center_suitability_notes": notes,
            "sources": sources,
            "researched_at": datetime.now(timezone.utc).isoformat(),
        },
    }


# ── Strategy 1: WARN Act Scraper ─────────────────────────────────────────


# State WARN Act data sources — URLs and formats
# Updated March 2026 with verified working URLs
WARN_SOURCES = {
    "AL": {
        "url": "https://workforce.alabama.gov/documents/warn-list/",
        "format": "csv",
        "scraper": "scrape_warn_al",
        "notes": "Alabama CSV feed — cleanest source, 1998-present",
    },
    "TX": {
        "url": "https://data.texas.gov/resource/8w53-c4f6.csv",
        "format": "csv",
        "scraper": "scrape_warn_tx",
        "notes": "Texas Socrata open data API — CSV export",
    },
    "GA": {
        "url": "https://www.tcsg.edu/warn-public-view/",
        "format": "html",
        "scraper": "scrape_warn_generic_html",
        "notes": "Georgia TCSG DataTables page — 2023+",
    },
    "IN": {
        "url": "https://www.in.gov/dwd/warn-notices/current-warn-notices/",
        "format": "html",
        "scraper": "scrape_warn_generic_html",
        "notes": "Indiana DWD HTML page",
    },
    "PA": {
        "url": "https://www.dli.pa.gov/Individuals/Workforce-Development/warn/notices/Pages/default.aspx",
        "format": "html",
        "scraper": "scrape_warn_generic_html",
        "notes": "Pennsylvania L&I HTML page",
    },
    "TN": {
        "url": "https://www.tn.gov/workforce/general-resources/major-publications0/major-publications-redirect/reports.html",
        "format": "html",
        "scraper": "scrape_warn_tn",
        "notes": "Tennessee HTML table with company, county, affected workers",
    },
    "VA": {
        "url": "https://www.vec.virginia.gov/warn-notices",
        "format": "html",
        "scraper": "scrape_warn_generic_html",
        "notes": "Virginia Employment Commission WARN page",
    },
    "NJ": {
        "url": "https://www.nj.gov/labor/assets/PDFs/WARN/WARN_Notice_Archive.xlsx",
        "format": "excel",
        "scraper": "scrape_warn_nj",
        "notes": "New Jersey Excel workbook — multi-sheet, all years",
    },
}

# States that are PDF-only or require auth — flagged for manual review
WARN_SKIPPED = {
    "OH": "No centralized listing page — individual PDFs only (jfs.ohio.gov)",
    "NC": "PDF only — commerce.nc.gov workforce WARN reports",
    "SC": "PDF only — scworks.org layoff notification reports",
    "AZ": "PDF only — Arizona ICA",
    "NV": "Requires search form — DETR",
    "IA": "PDF only — Iowa Workforce Development",
    "OR": "PDF only — Oregon Employment Department",
    "WA": "PDF only — Washington ESD",
    "IL": "Monthly Excel downloads — illinoisworknet.com/LayoffRecovery (would need openpyxl + URL scraping)",
}


def fetch_url(url, timeout=30):
    """Fetch a URL and return the response body as text."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "GridSite-StrandedCapacity/1.0 (brian@gridsite.dev)"
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        raw = resp.read()
        try:
            return raw.decode("utf-8")
        except UnicodeDecodeError:
            return raw.decode("latin-1")


def fetch_bytes(url, timeout=30):
    """Fetch a URL and return raw bytes."""
    req = urllib.request.Request(url, headers={
        "User-Agent": "GridSite-StrandedCapacity/1.0 (brian@gridsite.dev)"
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def parse_warn_csv(text, state, url, company_col, city_col, employees_col,
                   type_col=None, date_col=None):
    """Generic CSV parser for WARN data. Applies 24-month date filter."""
    results = []
    reader = csv.DictReader(io.StringIO(text))
    total = 0
    skipped_date = 0
    for row in reader:
        total += 1
        company = row.get(company_col, "").strip()
        city = row.get(city_col, "").strip()
        employees_str = row.get(employees_col, "0").strip()
        notice_type = row.get(type_col, "") if type_col else ""
        notice_date = row.get(date_col, "") if date_col else ""

        # Date filter — skip records older than CUTOFF_MONTHS
        if notice_date and not is_within_cutoff(notice_date):
            skipped_date += 1
            continue

        employees = 0
        try:
            employees = int(re.sub(r"[^\d]", "", employees_str) or "0")
        except ValueError:
            pass

        combined = company + " " + notice_type
        if employees >= MIN_EMPLOYEES and matches_industry(combined):
            is_closure = any(w in (notice_type + " " + company).lower()
                            for w in ["clos", "shutdown", "permanent"])
            if not is_closure and "layoff" in notice_type.lower():
                if employees < 500:
                    continue

            location = ", ".join(filter(None, [city, state]))
            results.append({
                "company": company,
                "location": location,
                "city": city,
                "state": state,
                "employees": employees,
                "notice_type": notice_type.strip() if notice_type else "WARN Filing",
                "notice_date": normalize_date(notice_date) if notice_date else "",
                "is_closure": is_closure,
            })

    log_scrape("warn_act", state, url, total, len(results))
    date_note = " ({} skipped by date)".format(skipped_date) if skipped_date else ""
    print("    {}: {} total records, {} matched{}".format(state, total, len(results), date_note))
    return results


def scrape_warn_al():
    """Scrape Alabama WARN Act CSV feed."""
    url = WARN_SOURCES["AL"]["url"]
    try:
        text = fetch_url(url)
        # Alabama CSV columns vary — try common column names
        reader = csv.DictReader(io.StringIO(text))
        fields = reader.fieldnames or []
        # Find the right column names by inspecting headers
        company_col = next((f for f in fields if "company" in f.lower() or "name" in f.lower()), fields[4] if len(fields) > 4 else "")
        city_col = next((f for f in fields if "location" in f.lower() or "city" in f.lower()), fields[5] if len(fields) > 5 else "")
        emp_col = next((f for f in fields if "affected" in f.lower() or "worker" in f.lower() or "employee" in f.lower()), fields[6] if len(fields) > 6 else "")
        type_col = next((f for f in fields if "type" in f.lower()), fields[1] if len(fields) > 1 else None)
        date_col = next((f for f in fields if "date" in f.lower() or "announcement" in f.lower()), fields[2] if len(fields) > 2 else None)

        # Re-parse with detected columns
        return parse_warn_csv(text, "AL", url, company_col, city_col, emp_col, type_col, date_col)
    except Exception as e:
        print("    AL: Error — {}".format(e))
        log_scrape("warn_act", "AL", url, 0, 0, "Error: " + str(e))
        return []


def scrape_warn_ga():
    """Scrape Georgia WARN data via TCSG DataTables AJAX endpoint.

    Uses GravityView DataTables server-side processing endpoint.
    Columns: GA WARN ID, Company Name, Submitted Date, Total Affected Employees
    """
    ajax_url = "https://www.tcsg.edu/wp-admin/admin-ajax.php"
    results = []
    try:
        # POST to the DataTables AJAX endpoint
        params = urllib.parse.urlencode({
            "action": "gv_datatables_data",
            "view_id": "77460",
            "post_id": "77462",
            "draw": "1",
            "start": "0",
            "length": "500",  # request up to 500 records
        }).encode("utf-8")

        req = urllib.request.Request(ajax_url, data=params, headers={
            "User-Agent": "GridSite-StrandedCapacity/1.0",
            "Content-Type": "application/x-www-form-urlencoded",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://www.tcsg.edu/warn-public-view/",
        })

        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        records = data.get("data", [])
        total = len(records)

        for row in records:
            # Each row is a list of HTML cell values
            if not isinstance(row, list) or len(row) < 4:
                continue

            clean = [re.sub(r"<[^>]+>", "", str(c)).strip() for c in row]
            # [0]=GA WARN ID, [1]=Company Name, [2]=Submitted Date, [3]=Total Affected, [4]=Entry ID
            company = clean[1] if len(clean) > 1 else ""
            date_str = clean[2] if len(clean) > 2 else ""
            emp_str = clean[3] if len(clean) > 3 else "0"

            employees = 0
            try:
                employees = int(re.sub(r"[^\d]", "", emp_str) or "0")
            except ValueError:
                pass

            if employees >= MIN_EMPLOYEES and matches_industry(company):
                is_closure = "clos" in company.lower() or "shut" in company.lower()
                if not is_closure and employees < 500:
                    continue

                results.append({
                    "company": company,
                    "location": "GA",
                    "city": "",
                    "state": "GA",
                    "employees": employees,
                    "notice_type": "WARN Filing",
                    "notice_date": date_str,
                    "is_closure": is_closure,
                })

        log_scrape("warn_act", "GA", ajax_url, total, len(results))
        print("    GA: {} DataTables records, {} matched".format(total, len(results)))

    except Exception as e:
        print("    GA: Error — {}".format(e))
        log_scrape("warn_act", "GA", ajax_url, 0, 0, "Error: " + str(e))

    return results


def scrape_warn_tx():
    """Scrape Texas WARN Act data via Socrata open data API.

    Columns: notice_date, job_site_name, county_name, wda_name,
             total_layoff_number, layoff_date, wfdd_received_date, city_name
    """
    base_url = WARN_SOURCES["TX"]["url"]
    query_url = base_url + "?$limit=5000&$order=notice_date%20DESC"
    try:
        text = fetch_url(query_url)
        return parse_warn_csv(
            text, "TX", query_url,
            company_col="job_site_name",
            city_col="city_name",
            employees_col="total_layoff_number",
            type_col=None,
            date_col="notice_date",
        )
    except Exception as e:
        print("    TX: Error — {}".format(e))
        log_scrape("warn_act", "TX", base_url, 0, 0, "Error: " + str(e))
        return []


def scrape_warn_tn():
    """Scrape Tennessee WARN Act HTML table.

    Columns: Date of Posting, Company (link), County, Affected Workers,
             Closure/Layoff Date, Notice/Type
    """
    url = WARN_SOURCES["TN"]["url"]
    results = []
    try:
        text = fetch_url(url, timeout=45)

        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", text, re.DOTALL | re.IGNORECASE)
        total = 0
        for row_html in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.DOTALL | re.IGNORECASE)
            if len(cells) < 4:
                continue
            total += 1

            clean = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]

            # [0]=date, [1]=company, [2]=county, [3]=affected workers, [4]=closure date, [5]=notice
            notice_date = clean[0] if len(clean) > 0 else ""
            company = clean[1] if len(clean) > 1 else ""
            county = clean[2] if len(clean) > 2 else ""
            workers_str = clean[3] if len(clean) > 3 else "0"
            closure_date = clean[4] if len(clean) > 4 else ""
            notice_type = clean[5] if len(clean) > 5 else ""

            employees = 0
            try:
                employees = int(re.sub(r"[^\d]", "", workers_str) or "0")
            except ValueError:
                pass

            combined = company + " " + notice_type + " " + county
            if employees >= MIN_EMPLOYEES and matches_industry(combined):
                is_closure = any(w in (combined + " " + closure_date).lower()
                                for w in ["clos", "shutdown", "permanent"])
                if not is_closure and employees < 500:
                    continue

                results.append({
                    "company": company,
                    "location": county + ", TN" if county else "TN",
                    "city": county,
                    "state": "TN",
                    "employees": employees,
                    "notice_type": notice_type or "WARN Filing",
                    "notice_date": notice_date,
                    "is_closure": is_closure,
                })

        log_scrape("warn_act", "TN", url, total, len(results))
        print("    TN: {} table rows, {} matched".format(total, len(results)))

    except Exception as e:
        print("    TN: Error — {}".format(e))
        log_scrape("warn_act", "TN", url, 0, 0, "Error: " + str(e))

    return results


def scrape_warn_nj():
    """Scrape New Jersey WARN Act Excel workbook."""
    url = WARN_SOURCES["NJ"]["url"]
    results = []
    try:
        import openpyxl

        data = fetch_bytes(url, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        total = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    header = [str(c or "").strip().lower() for c in row]
                    continue
                if not header:
                    continue
                total += 1

                # Map columns
                vals = {}
                for i, h in enumerate(header):
                    if i < len(row):
                        vals[h] = str(row[i] or "").strip()

                company = vals.get("company", vals.get("company name", ""))
                city = vals.get("city", vals.get("location", ""))
                emp_str = vals.get("workforce affected", vals.get("employees", vals.get("# affected", "0")))
                date_str = vals.get("effective date", vals.get("date", vals.get("month posted", "")))

                employees = 0
                try:
                    employees = int(re.sub(r"[^\d]", "", emp_str) or "0")
                except ValueError:
                    pass

                # Date filter
                if date_str and not is_within_cutoff(date_str):
                    continue

                if employees >= MIN_EMPLOYEES and matches_industry(company):
                    line = " ".join(vals.values())
                    is_closure = any(w in line.lower() for w in ["clos", "shutdown", "permanent"])
                    if not is_closure and employees < 500:
                        continue

                    results.append({
                        "company": company,
                        "location": city + ", NJ" if city else "NJ",
                        "city": city,
                        "state": "NJ",
                        "employees": employees,
                        "notice_type": "WARN Filing",
                        "notice_date": normalize_date(date_str),
                        "is_closure": is_closure,
                    })

        wb.close()
        log_scrape("warn_act", "NJ", url, total, len(results))
        print("    NJ: {} total records across {} sheets, {} matched".format(
            total, len(wb.sheetnames), len(results)))

    except ImportError:
        print("    NJ: SKIP — openpyxl not installed (pip install openpyxl)")
        log_scrape("warn_act", "NJ", url, 0, 0, "openpyxl not installed")
    except Exception as e:
        print("    NJ: Error — {}".format(e))
        log_scrape("warn_act", "NJ", url, 0, 0, "Error: " + str(e))

    return results


def scrape_warn_in():
    """Scrape Indiana WARN Act HTML table.

    Columns: Company, City, Employees, Notice Date, Effective Date, NAICS, Industry, Type (CL/LO)
    """
    url = WARN_SOURCES["IN"]["url"]
    results = []
    try:
        text = fetch_url(url, timeout=30)

        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", text, re.DOTALL | re.IGNORECASE)
        total = 0
        for row_html in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.DOTALL | re.IGNORECASE)
            if len(cells) < 5:
                continue
            total += 1

            clean = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            # [0]=Company, [1]=City, [2]=Employees, [3]=Notice Date, [4]=Effective Date,
            # [5]=NAICS, [6]=Industry, [7]=Type (CL/LO)
            company = clean[0]
            city = clean[1] if len(clean) > 1 else ""
            emp_str = clean[2] if len(clean) > 2 else "0"
            notice_date = clean[3] if len(clean) > 3 else ""
            effective_date = clean[4] if len(clean) > 4 else ""
            industry = clean[6] if len(clean) > 6 else ""
            warn_type = clean[7].strip().upper() if len(clean) > 7 else ""

            employees = 0
            try:
                employees = int(re.sub(r"[^\d]", "", emp_str) or "0")
            except ValueError:
                pass

            # Use effective date for cutoff check (when closure actually happens)
            date_for_filter = effective_date or notice_date
            if not is_within_cutoff(date_for_filter):
                continue

            combined = company + " " + industry
            if employees >= MIN_EMPLOYEES and matches_industry(combined):
                # CL = Closure, LO = Layoff
                is_closure = warn_type == "CL" or "clos" in combined.lower()
                if not is_closure and employees < 500:
                    continue

                results.append({
                    "company": company,
                    "location": city + ", IN" if city else "IN",
                    "city": city,
                    "state": "IN",
                    "employees": employees,
                    "notice_type": "Closure" if is_closure else "Layoff",
                    "notice_date": normalize_date(effective_date or notice_date),
                    "is_closure": is_closure,
                })

        log_scrape("warn_act", "IN", url, total, len(results))
        print("    IN: {} table rows, {} matched (after {}mo date filter)".format(
            total, len(results), CUTOFF_MONTHS))

    except Exception as e:
        print("    IN: Error — {}".format(e))
        log_scrape("warn_act", "IN", url, 0, 0, "Error: " + str(e))

    return results


def scrape_warn_generic_html(state):
    """Scrape WARN data from HTML page with table extraction."""
    info = WARN_SOURCES.get(state)
    if not info:
        return []

    url = info["url"]
    results = []
    try:
        text = fetch_url(url, timeout=30)

        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", text, re.DOTALL | re.IGNORECASE)
        total = 0
        for row_html in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, re.DOTALL | re.IGNORECASE)
            if len(cells) < 3:
                continue
            total += 1

            clean = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            line = " ".join(clean)

            if matches_industry(line):
                employees = 0
                for c in clean:
                    nums = re.findall(r"\d[\d,]*", c)
                    for n in nums:
                        v = int(n.replace(",", ""))
                        if 50 <= v <= 50000:
                            employees = max(employees, v)

                if employees >= MIN_EMPLOYEES:
                    company = clean[0] if clean else "Unknown"
                    location = ""
                    for c in clean[1:4]:
                        if any(c.upper() == s for s in TARGET_STATES + ["NJ", "AL"]):
                            continue
                        if len(c) > 3 and not c.isdigit():
                            location = c
                            break

                    results.append({
                        "company": company,
                        "location": location + ", " + state if location else state,
                        "city": location,
                        "state": state,
                        "employees": employees,
                        "notice_type": "WARN Filing",
                        "notice_date": "",
                        "is_closure": "clos" in line.lower() or "shut" in line.lower(),
                    })

        log_scrape("warn_act", state, url, total, len(results))
        print("    {}: {} table rows, {} matched".format(state, total, len(results)))

    except Exception as e:
        print("    {}: Error — {}".format(state, e))
        log_scrape("warn_act", state, url, 0, 0, "Error: " + str(e))

    return results


def run_warn_strategy(states, dry_run=False):
    """Run WARN Act scraper across target states."""
    print()
    print("=" * 70)
    print("STRATEGY 1: WARN Act Scraper")
    print("=" * 70)

    # Report skipped states
    for state in states:
        if state in WARN_SKIPPED:
            print("  SKIP {}: {}".format(state, WARN_SKIPPED[state]))

    if dry_run:
        active = [s for s in states if s in WARN_SOURCES]
        print("  Would scrape: {}".format(", ".join(active)))
        return []

    all_results = []
    print()
    print("  Scraping WARN Act data...")

    # Dispatch table for state-specific scrapers
    scrapers = {
        "AL": scrape_warn_al,
        "TX": scrape_warn_tx,
        "TN": scrape_warn_tn,
        "NJ": scrape_warn_nj,
        "GA": scrape_warn_ga,
        "IN": scrape_warn_in,
    }

    for state in states:
        if state not in WARN_SOURCES:
            continue

        if state in scrapers:
            results = scrapers[state]()
        else:
            results = scrape_warn_generic_html(state)

        all_results.extend(results)

    print()
    print("  Total WARN matches: {}".format(len(all_results)))

    # Geocode and build features — use facility address lookup for precision
    features = []
    has_api = bool(os.environ.get("ANTHROPIC_API_KEY"))
    print("  Geocoding {} sites{}...".format(
        len(all_results),
        " (with address lookup)" if has_api else " (city-level only, no API key)"))

    for r in all_results:
        city = r.get("city", "") or r.get("location", "")
        if not city:
            continue

        sub_type = classify_sub_type(r["company"])
        est_mw = estimate_mw(r["company"])
        if est_mw < MIN_MW:
            continue

        # Try facility address lookup via API first
        approximate = True
        coords = None
        if has_api:
            facility_addr, approx = lookup_facility_address(
                r["company"], city, r["state"])
            if facility_addr:
                coords = geocode_cached(facility_addr, r["state"])
                if coords:
                    approximate = approx
                    print("    {} — {} (precise: {})".format(
                        r["company"][:35], facility_addr[:40],
                        "yes" if not approx else "no"))

        # Fallback to city-level geocoding
        if not coords:
            coords = geocode_cached(city, r["state"])
            approximate = True

        if not coords:
            print("    SKIP (no geocode): {}".format(r["company"]))
            continue

        location_str = r["location"]
        if not approximate and has_api:
            # Use the precise address as location
            location_str = facility_addr or location_str

        feat = build_feature(
            name=r["company"],
            lat=coords[0], lon=coords[1],
            state=r["state"],
            source="warn_act",
            sub_type=sub_type,
            company=r["company"],
            location=location_str,
            estimated_mw=est_mw,
            employee_count=r["employees"],
            closure_date=r["notice_date"],
            closure_status="closing" if r["is_closure"] else "announced",
            sources=[WARN_SOURCES[r["state"]]["url"]],
            notes="WARN Act filing. {} employees affected.".format(r["employees"]),
            location_approximate=approximate,
        )
        features.append(feat)
        tag = "~" if approximate else ""
        print("    {}{} — {} ({}, {} emp, ~{} MW)".format(
            tag, r["company"][:38], r["state"], sub_type, r["employees"], est_mw))

    return features


# ── Strategy 2: EIA Large Consumer Analysis ──────────────────────────────


def run_eia_strategy(states, dry_run=False):
    """Research stranded capacity in utility territories using Anthropic API."""
    print()
    print("=" * 70)
    print("STRATEGY 2: EIA Large Consumer Analysis")
    print("=" * 70)

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("  SKIP — ANTHROPIC_API_KEY not set")
        return {}

    import anthropic

    # Load utility territories
    terr_file = os.path.join(DATA_DIR, "utility-territories.geojson")
    if not os.path.exists(terr_file):
        print("  SKIP — utility-territories.geojson not found")
        return {}

    with open(terr_file) as f:
        terr_geo = json.load(f)

    # Filter to target states and major utilities
    utilities = []
    seen = set()
    for feat in terr_geo["features"]:
        p = feat["properties"]
        state = p.get("state", "")
        name = p.get("name", "")
        if state in states and name and name not in seen:
            seen.add(name)
            utilities.append({"name": name, "state": state})

    print("  Utilities in target states: {}".format(len(utilities)))

    if dry_run:
        for u in utilities[:10]:
            print("    Would research: {} ({})".format(u["name"], u["state"]))
        if len(utilities) > 10:
            print("    ... and {} more".format(len(utilities) - 10))
        return {}

    client = anthropic.Anthropic()
    results = {}
    stats = {"researched": 0, "input_tokens": 0, "output_tokens": 0, "cost": 0}
    conn = sqlite3.connect(DB_FILE)

    for i, util in enumerate(utilities):
        site_key = "eia|{}|{}".format(util["name"], util["state"])

        # Check cache
        cur = conn.execute(
            "SELECT response_json FROM research_log WHERE site_key = ?", (site_key,))
        cached = cur.fetchone()
        if cached:
            results[site_key] = json.loads(cached[0])
            print("    CACHED {} ({})".format(util["name"][:40], util["state"]))
            continue

        prompt = (
            "You are an energy market research analyst. Research the following utility territory "
            "and return ONLY a JSON object. No other text.\n\n"
            "Utility: {}\n"
            "State: {}\n\n"
            "Research:\n"
            "1. Has this utility lost any major industrial customers (30MW+ load) in the last 3 years?\n"
            "2. Are there any large plant closures (manufacturing, steel, auto, chemical, paper, etc.) "
            "planned or recently completed in this utility's service territory?\n"
            "3. Has this utility mentioned stranded capacity, load loss, or declining industrial demand "
            "in any recent filings or press releases?\n\n"
            "Return ONLY this JSON:\n"
            '{{\n'
            '  "has_stranded_capacity": true | false,\n'
            '  "closures": [\n'
            '    {{\n'
            '      "facility_name": "...",\n'
            '      "company": "...",\n'
            '      "location": "city, state",\n'
            '      "estimated_mw": number,\n'
            '      "employee_count": number or null,\n'
            '      "closure_date": "YYYY-MM" or null,\n'
            '      "status": "closed" | "closing" | "announced",\n'
            '      "sub_type": "Steel Mill" | "Auto Assembly" | etc.,\n'
            '      "notes": "brief context"\n'
            '    }}\n'
            '  ],\n'
            '  "sources": ["url1", "url2"],\n'
            '  "summary": "one sentence summary"\n'
            '}}'
        ).format(util["name"], util["state"])

        try:
            response = client.messages.create(
                model=MODEL, max_tokens=MAX_TOKENS,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content": prompt}],
            )

            text = ""
            for block in response.content:
                if hasattr(block, "text"):
                    text += block.text

            input_t = response.usage.input_tokens
            output_t = response.usage.output_tokens
            cost = input_t * INPUT_COST_PER_TOKEN + output_t * OUTPUT_COST_PER_TOKEN
            stats["input_tokens"] += input_t
            stats["output_tokens"] += output_t
            stats["cost"] += cost
            stats["researched"] += 1

            # Parse response
            result = None
            text = text.strip()
            start = text.find("{")
            end = text.rfind("}") + 1
            if start >= 0 and end > start:
                try:
                    result = json.loads(text[start:end])
                except json.JSONDecodeError:
                    pass

            if result is None:
                result = {"has_stranded_capacity": False, "closures": [], "sources": [], "summary": "Parse error"}

            results[site_key] = result

            conn.execute(
                "INSERT OR REPLACE INTO research_log (site_key, strategy, site_name, state, response_json, input_tokens, output_tokens, cost_usd, researched_at, model) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (site_key, "eia_analysis", util["name"], util["state"],
                 json.dumps(result), input_t, output_t, cost,
                 datetime.now(timezone.utc).isoformat(), MODEL),
            )
            conn.commit()

            has = result.get("has_stranded_capacity", False)
            n_closures = len(result.get("closures", []))
            icon = "FOUND" if has else "NONE"
            print("    {:6s} {} ({}) — {} closures (${:.3f})".format(
                icon, util["name"][:35], util["state"], n_closures, cost))

        except Exception as e:
            print("    ERROR {} ({}) — {}".format(util["name"][:35], util["state"], e))

    conn.close()
    print()
    print("  EIA research: {} utilities, ${:.2f}".format(stats["researched"], stats["cost"]))
    return results


# ── Strategy 3: News & SEC Filing Scanner ─────────────────────────────────


NEWS_QUERIES = [
    "plant closure {state} 2025 2026 manufacturing",
    "factory closing {state} 2025 2026",
    "steel mill closure {state}",
    "paper mill closure {state}",
    "smelter closure {state}",
    "refinery closure {state}",
    "distribution center closing {state} 2025 2026",
    "auto assembly plant shutdown {state}",
]


def run_news_strategy(states, dry_run=False):
    """Scan news for industrial closures using Anthropic API with web search."""
    print()
    print("=" * 70)
    print("STRATEGY 3: News & SEC Filing Scanner")
    print("=" * 70)

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("  SKIP — ANTHROPIC_API_KEY not set")
        return []

    import anthropic

    if dry_run:
        total_queries = len(states) * 2  # 2 batched searches per state
        print("  Would run {} research queries across {} states".format(total_queries, len(states)))
        return []

    client = anthropic.Anthropic()
    all_closures = []
    stats = {"researched": 0, "input_tokens": 0, "output_tokens": 0, "cost": 0}
    conn = sqlite3.connect(DB_FILE)

    for state in states:
        site_key = "news|{}".format(state)

        # Check cache
        cur = conn.execute(
            "SELECT response_json FROM research_log WHERE site_key = ?", (site_key,))
        cached = cur.fetchone()
        if cached:
            result = json.loads(cached[0])
            closures = result.get("closures", [])
            for c in closures:
                c["state"] = state
            all_closures.extend(closures)
            print("    CACHED {} — {} closures".format(state, len(closures)))
            continue

        prompt = (
            "You are an industrial real estate research analyst. Search for recent and "
            "announced industrial facility closures in {} that might leave behind stranded "
            "power capacity suitable for data center conversion.\n\n"
            "Search for:\n"
            "- Manufacturing plant closures/shutdowns in {} (2024-2026)\n"
            "- Steel mill, paper mill, smelter, refinery closures in {}\n"
            "- Auto assembly, chemical plant closures in {}\n"
            "- Large distribution center/fulfillment center closures in {}\n"
            "- Any large industrial facility (200+ employees) closing in {}\n\n"
            "For each closure found, extract:\n"
            "- Facility name and company\n"
            "- City/location\n"
            "- Employee count\n"
            "- Estimated power consumption in MW (based on facility type)\n"
            "- Closure date (actual or planned)\n"
            "- Facility type (steel mill, auto assembly, paper mill, etc.)\n\n"
            "Return ONLY this JSON:\n"
            '{{\n'
            '  "closures": [\n'
            '    {{\n'
            '      "facility_name": "...",\n'
            '      "company": "...",\n'
            '      "city": "...",\n'
            '      "employee_count": number,\n'
            '      "estimated_mw": number,\n'
            '      "closure_date": "YYYY-MM or description",\n'
            '      "status": "closed" | "closing" | "announced",\n'
            '      "sub_type": "Steel Mill" | "Auto Assembly" | "Paper Mill" | etc.,\n'
            '      "notes": "brief context",\n'
            '      "sources": ["url1"]\n'
            '    }}\n'
            '  ]\n'
            '}}'
        ).format(state, state, state, state, state, state)

        try:
            response = client.messages.create(
                model=MODEL, max_tokens=MAX_TOKENS,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content": prompt}],
            )

            text = ""
            for block in response.content:
                if hasattr(block, "text"):
                    text += block.text

            input_t = response.usage.input_tokens
            output_t = response.usage.output_tokens
            cost = input_t * INPUT_COST_PER_TOKEN + output_t * OUTPUT_COST_PER_TOKEN
            stats["input_tokens"] += input_t
            stats["output_tokens"] += output_t
            stats["cost"] += cost
            stats["researched"] += 1

            result = None
            start = text.find("{")
            end = text.rfind("}") + 1
            if start >= 0 and end > start:
                try:
                    result = json.loads(text[start:end])
                except json.JSONDecodeError:
                    pass

            if result is None:
                result = {"closures": []}

            closures = result.get("closures", [])
            for c in closures:
                c["state"] = state

            conn.execute(
                "INSERT OR REPLACE INTO research_log (site_key, strategy, site_name, state, response_json, input_tokens, output_tokens, cost_usd, researched_at, model) VALUES (?,?,?,?,?,?,?,?,?,?)",
                (site_key, "news_scan", state, state, json.dumps(result),
                 input_t, output_t, cost,
                 datetime.now(timezone.utc).isoformat(), MODEL),
            )
            conn.commit()

            # Filter to >= MIN_MW
            qualified = [c for c in closures if c.get("estimated_mw", 0) >= MIN_MW]
            all_closures.extend(qualified)
            print("    {} — {} closures found, {} >= {}MW (${:.3f})".format(
                state, len(closures), len(qualified), MIN_MW, cost))

        except Exception as e:
            print("    {} — ERROR: {}".format(state, e))

    conn.close()

    # Geocode and build features
    features = []
    print()
    print("  Geocoding {} closure sites...".format(len(all_closures)))
    for c in all_closures:
        city = c.get("city", "")
        state = c.get("state", "")
        if not city:
            continue

        coords = geocode_cached(city, state)
        if not coords:
            print("    SKIP (no geocode): {} in {}".format(c.get("facility_name", "?"), city))
            continue

        feat = build_feature(
            name=c.get("facility_name", c.get("company", "Unknown")),
            lat=coords[0], lon=coords[1],
            state=state,
            source="news_scan",
            sub_type=c.get("sub_type", classify_sub_type(c.get("facility_name", ""))),
            company=c.get("company", ""),
            location="{}, {}".format(city, state),
            estimated_mw=c.get("estimated_mw", 30),
            employee_count=c.get("employee_count", 0),
            closure_date=c.get("closure_date", ""),
            closure_status=c.get("status", "announced"),
            sources=c.get("sources", []),
            notes=c.get("notes", ""),
        )
        features.append(feat)
        print("    {} — {}, {} (~{} MW, {} emp)".format(
            c.get("facility_name", "?")[:35], city, state,
            c.get("estimated_mw", "?"), c.get("employee_count", "?")))

    print()
    print("  News scan: {} states, {} features, ${:.2f}".format(
        stats["researched"], len(features), stats["cost"]))
    return features


# ── Main ──────────────────────────────────────────────────────────────────


def write_geojson(features, output_path):
    geo = {"type": "FeatureCollection", "features": features}
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(geo, f, indent=2)
    size = round(os.path.getsize(output_path) / 1024, 1)
    print("  Output: {} ({} KB, {} features)".format(output_path, size, len(features)))


def main():
    parser = argparse.ArgumentParser(description="Stranded Capacity Pipeline")
    parser.add_argument("--warn-only", action="store_true", help="Run WARN Act scraper only")
    parser.add_argument("--news-only", action="store_true", help="Run news scanner only")
    parser.add_argument("--eia-only", action="store_true", help="Run EIA analysis only")
    parser.add_argument("--states", type=str, help="Comma-separated state codes")
    parser.add_argument("--dry-run", action="store_true", help="Show plan without executing")
    args = parser.parse_args()

    states = args.states.split(",") if args.states else TARGET_STATES
    run_all = not (args.warn_only or args.news_only or args.eia_only)

    print("=" * 70)
    print("STRANDED CAPACITY PIPELINE")
    print("=" * 70)
    print("  States: {}".format(", ".join(states)))
    print("  Strategies: {}".format(
        "ALL" if run_all else
        "WARN" if args.warn_only else
        "NEWS" if args.news_only else "EIA"
    ))

    init_db()

    # Load infrastructure data early
    print("  Loading infrastructure data...")
    subs = load_substations()
    print("    Substations >= 138kV: {:,}".format(len(subs)))

    warn_features = []
    news_features = []
    eia_results = {}

    if run_all or args.warn_only:
        warn_features = run_warn_strategy(states, args.dry_run)
        if warn_features:
            write_geojson(warn_features, WARN_OUTPUT)

    if run_all or args.news_only:
        news_features = run_news_strategy(states, args.dry_run)
        if news_features:
            write_geojson(news_features, NEWS_OUTPUT)

    if run_all or args.eia_only:
        eia_results = run_eia_strategy(states, args.dry_run)
        if eia_results:
            # Extract closures from EIA results into features
            eia_features = []
            for key, result in eia_results.items():
                closures = result.get("closures", [])
                for c in closures:
                    if c.get("estimated_mw", 0) < MIN_MW:
                        continue
                    city = c.get("location", "")
                    state = key.split("|")[-1] if "|" in key else ""
                    if city:
                        coords = geocode_cached(city.split(",")[0].strip(), state)
                        if coords:
                            feat = build_feature(
                                name=c.get("facility_name", "Unknown"),
                                lat=coords[0], lon=coords[1],
                                state=state, source="eia_analysis",
                                sub_type=c.get("sub_type", "Industrial Facility"),
                                company=c.get("company", ""),
                                location=city,
                                estimated_mw=c.get("estimated_mw", 30),
                                employee_count=c.get("employee_count", 0),
                                closure_date=c.get("closure_date", ""),
                                closure_status=c.get("status", "announced"),
                                sources=result.get("sources", []),
                                notes=c.get("notes", ""),
                            )
                            eia_features.append(feat)

            # Save EIA research data
            with open(EIA_OUTPUT, "w") as f:
                json.dump({
                    "metadata": {
                        "generated_at": datetime.now(timezone.utc).isoformat(),
                        "utilities_researched": len(eia_results),
                    },
                    "results": eia_results,
                }, f, indent=2)
            print("  EIA research: {}".format(EIA_OUTPUT))

    # Summary
    print()
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print("  WARN Act sites:      {}".format(len(warn_features)))
    print("  News scan sites:     {}".format(len(news_features)))
    print("  EIA utilities:       {}".format(len(eia_results)))
    total = len(warn_features) + len(news_features)
    print("  Total new sites:     {}".format(total))
    print("  Database: {}".format(DB_FILE))


if __name__ == "__main__":
    main()
