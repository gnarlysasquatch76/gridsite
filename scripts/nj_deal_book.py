#!/usr/bin/env python3
"""
NJ Deal Book Pipeline — identify and score data center sites in New Jersey.

Runs a 7-part analysis:
  Part 1: NJ industrial closure scan (WARN Act + news + utility load loss)
  Part 2: NJ retired/retiring power plant analysis
  Part 3: NJ substation proximity + vacant industrial parcels
  Part 4: Transportation corridor overlay
  Part 5: Scoring and ranking (top 10)
  Part 6: Site briefs
  Part 7: Compile deal book (PDF + markdown)

Usage:
    python3 scripts/nj_deal_book.py                # Run all parts
    python3 scripts/nj_deal_book.py --part 1       # Run a specific part
    python3 scripts/nj_deal_book.py --dry-run      # Preview without API calls
    python3 scripts/nj_deal_book.py --budget        # Show current API spend
"""

import argparse
import json
import math
import os
import re
import sqlite3
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.join(SCRIPT_DIR, "..")
DATA_DIR = os.path.join(PROJECT_DIR, "public", "data")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "output", "nj-deal-book")
DB_FILE = os.path.join(PROJECT_DIR, "nj_deal_book.db")

# Existing data files
SUBSTATIONS_FILE = os.path.join(DATA_DIR, "substations.geojson")
TRANSMISSION_FILE = os.path.join(DATA_DIR, "transmission-lines.geojson")
PLANTS_FILE = os.path.join(DATA_DIR, "power-plants.geojson")

# Output files
CLOSURES_FILE = os.path.join(OUTPUT_DIR, "nj-closures.json")
PLANTS_NJ_FILE = os.path.join(OUTPUT_DIR, "nj-plants.json")
SUBSTATIONS_NJ_FILE = os.path.join(OUTPUT_DIR, "nj-substations.json")
PARCELS_FILE = os.path.join(OUTPUT_DIR, "nj-parcels.json")
ALL_SITES_FILE = os.path.join(OUTPUT_DIR, "nj-all-sites.json")
SCORED_FILE = os.path.join(OUTPUT_DIR, "nj-scored-sites.json")
BRIEFS_FILE = os.path.join(OUTPUT_DIR, "nj-site-briefs.json")
DEALBOOK_MD = os.path.join(OUTPUT_DIR, "nj-deal-book.md")
DEALBOOK_PDF = os.path.join(OUTPUT_DIR, "nj-deal-book.pdf")

# Anthropic API settings — use Haiku for budget efficiency
MODEL = "claude-sonnet-4-20250514"
MAX_TOKENS = 2000
INPUT_COST_PER_TOKEN = 3.0 / 1_000_000
OUTPUT_COST_PER_TOKEN = 15.0 / 1_000_000
BUDGET_LIMIT = 5.00  # $5 max spend

# NJ-specific constants
NJ_WARN_URL = "https://www.nj.gov/labor/assets/PDFs/WARN/WARN_Notice_Archive.xlsx"
CUTOFF_MONTHS = 36  # Spec says 36 months for NJ deal book
MIN_EMPLOYEES = 100  # Spec says 100+ (lower than nationwide 200)
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
NOMINATIM_DELAY = 1.1

# Industry keywords for NJ (expanded per spec)
NJ_INDUSTRY_KEYWORDS = [
    "manufactur", "chemical", "pharmaceutical", "pharma", "steel", "glass",
    "automotive", "auto", "distribution", "fulfillment", "food process",
    "data center", "warehouse", "logistics", "refinery", "paper", "mill",
    "smelter", "aluminum", "cement", "foundry", "semiconductor", "plastics",
    "metals", "fabricat", "processing", "mining", "assembly", "plant",
]

# Power draw heuristic (from spec)
MW_ESTIMATES = {
    "data center": 50, "steel": 50, "glass": 50, "smelter": 60,
    "aluminum": 60, "chemical": 25, "pharmaceutical": 25, "pharma": 25,
    "refinery": 30, "paper": 20, "cement": 30, "foundry": 20,
    "semiconductor": 30, "auto": 15, "automotive": 15, "assembly": 15,
    "distribution": 10, "fulfillment": 12, "warehouse": 8,
    "food process": 10, "manufacturing": 10, "processing": 10,
    "logistics": 8, "mill": 15, "plastics": 12, "metals": 15,
    "fabricat": 10, "mining": 20, "plant": 10,
}

# Scale MW estimate by employee count per spec heuristic
def estimate_mw(text, employees=0):
    """Estimate MW from facility type and employee count."""
    text_lower = text.lower()
    base_mw = 10
    for keyword, mw in sorted(MW_ESTIMATES.items(), key=lambda x: -x[1]):
        if keyword in text_lower:
            base_mw = mw
            break

    # Scale by employee count using spec ranges
    if employees >= 500:
        scale = 2.0
    elif employees >= 300:
        scale = 1.5
    elif employees >= 200:
        scale = 1.2
    elif employees >= 100:
        scale = 1.0
    else:
        scale = 0.8

    return round(base_mw * scale)


def classify_sub_type(text):
    """Classify facility sub-type from description."""
    text_lower = text.lower()
    mappings = [
        ("steel", "Steel Mill"), ("smelter", "Smelter"), ("aluminum", "Aluminum Smelter"),
        ("auto", "Auto Assembly"), ("paper", "Paper Mill"), ("chemical", "Chemical Plant"),
        ("pharma", "Pharmaceutical"), ("refiner", "Refinery"), ("glass", "Glass Plant"),
        ("cement", "Cement Plant"), ("foundry", "Foundry"), ("semiconductor", "Semiconductor Fab"),
        ("data center", "Data Center"), ("distribution", "Distribution Center"),
        ("fulfillment", "Fulfillment Center"), ("warehouse", "Warehouse/Logistics"),
        ("logistics", "Warehouse/Logistics"), ("food", "Food Processing"),
        ("mining", "Mining Operation"), ("mill", "Industrial Mill"),
        ("processing", "Processing Plant"), ("manufactur", "Manufacturing"),
    ]
    for keyword, label in mappings:
        if keyword in text_lower:
            return label
    return "Industrial Facility"


def matches_industry(text):
    text_lower = text.lower()
    return any(kw in text_lower for kw in NJ_INDUSTRY_KEYWORDS)


# ── Utilities ────────────────────────────────────────────────────────────────


def haversine_miles(lat1, lon1, lat2, lon2):
    R = 3958.8
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat / 2) ** 2 +
         math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) *
         math.sin(dlon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def parse_date(date_str):
    if not date_str or not date_str.strip():
        return None
    date_str = date_str.strip()
    for fmt in [
        "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
        "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
        "%B %d, %Y", "%b %d, %Y", "%d-%b-%Y",
    ]:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def is_within_cutoff(date_str):
    dt = parse_date(date_str)
    if dt is None:
        return True
    cutoff = datetime.now() - timedelta(days=CUTOFF_MONTHS * 30)
    return dt >= cutoff


def normalize_date(date_str):
    dt = parse_date(date_str)
    if dt is None:
        return date_str.strip() if date_str else ""
    return dt.strftime("%Y-%m-%d")


def fetch_bytes(url, timeout=30):
    req = urllib.request.Request(url, headers={
        "User-Agent": "GridSite-NJDealBook/1.0 (brian@gridsite.dev)"
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def parse_json_response(text):
    """Extract JSON from an API response text."""
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    if "```" in text:
        for block in text.split("```"):
            block = block.strip()
            if block.startswith("json"):
                block = block[4:].strip()
            try:
                return json.loads(block)
            except json.JSONDecodeError:
                continue
    start = text.find("{")
    end = text.rfind("}") + 1
    if start >= 0 and end > start:
        try:
            return json.loads(text[start:end])
        except json.JSONDecodeError:
            pass
    # Try array
    start = text.find("[")
    end = text.rfind("]") + 1
    if start >= 0 and end > start:
        try:
            return json.loads(text[start:end])
        except json.JSONDecodeError:
            pass
    return None


# ── Database ─────────────────────────────────────────────────────────────────


def init_db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS api_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source TEXT,
            query_type TEXT,
            query TEXT,
            response_json TEXT,
            input_tokens INTEGER,
            output_tokens INTEGER,
            cost_usd REAL,
            timestamp TEXT,
            model TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cache (
            cache_key TEXT PRIMARY KEY,
            response_json TEXT,
            cached_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS geocode_cache (
            address TEXT PRIMARY KEY,
            lat REAL,
            lon REAL,
            cached_at TEXT
        )
    """)
    conn.commit()
    conn.close()


def get_total_spend():
    """Get total API spend from this pipeline."""
    if not os.path.exists(DB_FILE):
        return 0.0
    conn = sqlite3.connect(DB_FILE)
    cur = conn.execute("SELECT COALESCE(SUM(cost_usd), 0) FROM api_log")
    total = cur.fetchone()[0]
    conn.close()
    return total


def check_budget(estimated_cost=0.0):
    """Check if we're within budget. Returns (ok, remaining)."""
    spent = get_total_spend()
    remaining = BUDGET_LIMIT - spent
    if spent + estimated_cost > BUDGET_LIMIT:
        return False, remaining
    return True, remaining


def log_api_call(source, query_type, query, response, input_tokens, output_tokens, cost):
    """Log an API call to SQLite and print cost."""
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO api_log (source, query_type, query, response_json, input_tokens, output_tokens, cost_usd, timestamp, model) VALUES (?,?,?,?,?,?,?,?,?)",
        (source, query_type, query[:500], json.dumps(response) if isinstance(response, dict) else str(response)[:2000],
         input_tokens, output_tokens, cost,
         datetime.now(timezone.utc).isoformat(), MODEL),
    )
    conn.commit()
    conn.close()


def get_cached(cache_key, max_age_days=7):
    """Get a cached result if fresh enough."""
    conn = sqlite3.connect(DB_FILE)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=max_age_days)).isoformat()
    cur = conn.execute(
        "SELECT response_json FROM cache WHERE cache_key = ? AND cached_at > ?",
        (cache_key, cutoff))
    row = cur.fetchone()
    conn.close()
    if row:
        return json.loads(row[0])
    return None


def set_cached(cache_key, data):
    """Cache a result."""
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT OR REPLACE INTO cache (cache_key, response_json, cached_at) VALUES (?,?,?)",
        (cache_key, json.dumps(data), datetime.now(timezone.utc).isoformat()))
    conn.commit()
    conn.close()


# ── Geocoding ────────────────────────────────────────────────────────────────


def geocode(address, state=""):
    query = address
    if state and state not in address:
        query = address + ", " + state
    query = query + ", USA"
    params = urllib.parse.urlencode({
        "q": query, "format": "json", "limit": "1", "countrycodes": "us",
    })
    url = NOMINATIM_URL + "?" + params
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "GridSite-NJDealBook/1.0 (brian@gridsite.dev)"
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print("      Geocode failed for '{}': {}".format(address, e))
    return None


def geocode_cached(address, state=""):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.execute("SELECT lat, lon FROM geocode_cache WHERE address = ?", (address,))
    row = cur.fetchone()
    conn.close()
    if row:
        return row[0], row[1]

    time.sleep(NOMINATIM_DELAY)
    result = geocode(address, state)
    if result:
        conn = sqlite3.connect(DB_FILE)
        conn.execute(
            "INSERT OR REPLACE INTO geocode_cache (address, lat, lon, cached_at) VALUES (?,?,?,?)",
            (address, result[0], result[1], datetime.now(timezone.utc).isoformat()))
        conn.commit()
        conn.close()
    return result


# ── Infrastructure data ──────────────────────────────────────────────────────

_substations_nj = None
_all_substations = None


def load_nj_substations(min_kv=138):
    """Load NJ substations with >= min_kv voltage."""
    global _substations_nj
    if _substations_nj is not None:
        return _substations_nj
    _substations_nj = []
    if not os.path.exists(SUBSTATIONS_FILE):
        print("  WARNING: substations.geojson not found")
        return _substations_nj
    with open(SUBSTATIONS_FILE) as f:
        geo = json.load(f)
    for feat in geo["features"]:
        p = feat["properties"]
        v = p.get("MAX_VOLT")
        state = p.get("STATE", "")
        if v is not None and float(v) >= min_kv and state == "NJ":
            coords = feat["geometry"]["coordinates"]
            _substations_nj.append({
                "lat": float(p.get("LATITUDE", coords[1])),
                "lon": float(p.get("LONGITUDE", coords[0])),
                "max_volt": float(v),
                "name": p.get("NAME", ""),
                "city": p.get("CITY", ""),
                "county": p.get("COUNTY", ""),
            })
    return _substations_nj


def load_all_substations(min_kv=138):
    """Load all substations (for nearest-substation lookups)."""
    global _all_substations
    if _all_substations is not None:
        return _all_substations
    _all_substations = []
    if not os.path.exists(SUBSTATIONS_FILE):
        return _all_substations
    with open(SUBSTATIONS_FILE) as f:
        geo = json.load(f)
    for feat in geo["features"]:
        p = feat["properties"]
        v = p.get("MAX_VOLT")
        if v is not None and float(v) >= min_kv:
            coords = feat["geometry"]["coordinates"]
            _all_substations.append({
                "lat": float(p.get("LATITUDE", coords[1])),
                "lon": float(p.get("LONGITUDE", coords[0])),
                "max_volt": float(v),
                "name": p.get("NAME", ""),
                "state": p.get("STATE", ""),
            })
    return _all_substations


def find_nearest_substation(lat, lon, min_kv=138):
    """Find nearest substation. Returns dict or None."""
    subs = load_all_substations(min_kv)
    best_dist = float("inf")
    best = None
    for s in subs:
        d = haversine_miles(lat, lon, s["lat"], s["lon"])
        if d < best_dist:
            best_dist = d
            best = s
            best["distance_miles"] = round(d, 2)
    return best


# ── Anthropic API ────────────────────────────────────────────────────────────


def call_anthropic(prompt, source, query_type, cache_key=None, max_tokens=None):
    """Call Anthropic API with web search, caching, budget tracking, and Supabase logging."""
    if cache_key:
        cached = get_cached(cache_key)
        if cached:
            print("    [CACHED] {}".format(query_type))
            return cached

    ok, remaining = check_budget(0.05)  # estimate ~$0.05 per call
    if not ok:
        print("    [BUDGET] ${:.2f} remaining — skipping {}".format(remaining, query_type))
        return None

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("    [SKIP] No ANTHROPIC_API_KEY set")
        return None

    import anthropic
    client = anthropic.Anthropic()

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=max_tokens or MAX_TOKENS,
            tools=[{"type": "web_search_20250305", "name": "web_search"}],
            messages=[{"role": "user", "content": prompt}],
        )

        text = ""
        for block in response.content:
            if hasattr(block, "text"):
                text += block.text

        input_tokens = response.usage.input_tokens
        output_tokens = response.usage.output_tokens
        cost = input_tokens * INPUT_COST_PER_TOKEN + output_tokens * OUTPUT_COST_PER_TOKEN

        result = parse_json_response(text)
        if result is None:
            result = {"raw_text": text[:2000], "parse_error": True}

        log_api_call(source, query_type, prompt[:500], result, input_tokens, output_tokens, cost)

        if cache_key:
            set_cached(cache_key, result)

        spent = get_total_spend()
        print("    [API] {} — ${:.3f} (total: ${:.2f}/{:.2f})".format(
            query_type, cost, spent, BUDGET_LIMIT))

        return result

    except Exception as e:
        print("    [ERROR] {} — {}".format(query_type, e))
        return None


# ══════════════════════════════════════════════════════════════════════════════
# PART 1: NJ Industrial Closure Scan
# ══════════════════════════════════════════════════════════════════════════════


def scrape_nj_warn():
    """Scrape NJ WARN Act Excel workbook for plant closures, 100+ employees, 36 months."""
    print("\n  --- NJ WARN Act Scraper ---")
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl not installed (pip install openpyxl)")
        return []

    print("  Downloading NJ WARN Excel...")
    try:
        import io
        data = fetch_bytes(NJ_WARN_URL, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    except Exception as e:
        print("  ERROR downloading WARN data: {}".format(e))
        return []

    results = []
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                header = [str(c or "").strip().lower() for c in row]
                continue
            if not header:
                continue
            total += 1

            vals = {}
            for i, h in enumerate(header):
                if i < len(row):
                    vals[h] = str(row[i] or "").strip()

            company = vals.get("company", vals.get("company name", ""))
            city = vals.get("city", vals.get("location", ""))
            emp_str = vals.get("workforce affected",
                             vals.get("employees",
                             vals.get("# affected", "0")))
            date_str = vals.get("effective date",
                              vals.get("date",
                              vals.get("month posted", "")))

            employees = 0
            try:
                employees = int(re.sub(r"[^\d]", "", emp_str) or "0")
            except ValueError:
                pass

            if date_str and not is_within_cutoff(date_str):
                continue

            if employees >= MIN_EMPLOYEES and matches_industry(company):
                line = " ".join(vals.values())
                is_closure = any(w in line.lower() for w in ["clos", "shutdown", "permanent"])
                # Include all 100+ employee matches for NJ (spec is broader)
                if not is_closure and employees < 300:
                    continue

                results.append({
                    "company": company,
                    "location": city + ", NJ" if city else "NJ",
                    "city": city,
                    "state": "NJ",
                    "employees": employees,
                    "notice_date": normalize_date(date_str),
                    "is_closure": is_closure,
                    "source": "warn_act",
                })

    wb.close()
    print("  NJ WARN: {} total records, {} matched ({}+ emp, 36mo, industrial)".format(
        total, len(results), MIN_EMPLOYEES))
    return results


def run_nj_news_scan():
    """Scan for NJ industrial closures via Anthropic API with web search."""
    print("\n  --- NJ Industrial Closure News Scan ---")

    queries = [
        "major industrial closure New Jersey 2024 2025 2026",
        "factory closing New Jersey",
        "plant shutdown New Jersey PSEG JCP&L",
        "warehouse closing New Jersey large employer",
    ]

    all_closures = []

    for query in queries:
        cache_key = "news|NJ|{}".format(query[:50])

        prompt = (
            "You are an industrial real estate research analyst. Search for recent "
            "industrial facility closures in New Jersey matching this query:\n\n"
            '"{query}"\n\n'
            "Focus on:\n"
            "- Manufacturing plants, chemical/pharmaceutical plants, steel/glass facilities\n"
            "- Large distribution/fulfillment centers\n"
            "- Any facility with 100+ employees closing in NJ since January 2023\n"
            "- Data center closures or relocations\n\n"
            "For each closure found, extract details. Return ONLY this JSON:\n"
            '{{\n'
            '  "closures": [\n'
            '    {{\n'
            '      "facility_name": "...",\n'
            '      "company": "...",\n'
            '      "address": "full street address if available, or city, NJ",\n'
            '      "city": "...",\n'
            '      "county": "...",\n'
            '      "employee_count": number,\n'
            '      "estimated_mw": number,\n'
            '      "closure_date": "YYYY-MM or description",\n'
            '      "status": "closed" | "closing" | "announced",\n'
            '      "sub_type": "Manufacturing" | "Chemical Plant" | etc.,\n'
            '      "utility_territory": "PSEG" | "JCP&L" | "other",\n'
            '      "notes": "brief context",\n'
            '      "sources": ["url1"]\n'
            '    }}\n'
            '  ]\n'
            '}}'
        ).format(query=query)

        result = call_anthropic(prompt, "part1_news", "news_scan: " + query[:40], cache_key)
        if result and "closures" in result:
            for c in result["closures"]:
                c["state"] = "NJ"
                c["source"] = "news_scan"
            all_closures.extend(result["closures"])
        elif result and isinstance(result, list):
            for c in result:
                c["state"] = "NJ"
                c["source"] = "news_scan"
            all_closures.extend(result)

    print("  News scan: {} closures found across {} queries".format(
        len(all_closures), len(queries)))
    return all_closures


def run_nj_utility_scan():
    """Research PSEG and JCP&L load loss via Anthropic API."""
    print("\n  --- NJ Utility Load Loss Scan ---")

    queries = [
        ("PSEG lost major industrial customer New Jersey stranded capacity",
         "PSEG load loss"),
        ("JCP&L large customer closure New Jersey industrial",
         "JCP&L load loss"),
        ("New Jersey utility stranded capacity industrial load reduction 2024 2025",
         "NJ utility stranded"),
        ("PJM interconnection queue New Jersey withdrawn cancelled 2024 2025",
         "PJM queue withdrawals NJ"),
    ]

    all_findings = []

    for query, label in queries:
        cache_key = "utility|NJ|{}".format(label)

        prompt = (
            "You are an energy market research analyst. Research this topic:\n\n"
            '"{query}"\n\n'
            "Find specific instances of:\n"
            "- Large industrial customers (10MW+) that have left PSEG or JCP&L territory\n"
            "- Plant closures that freed up utility infrastructure in NJ\n"
            "- Withdrawn PJM interconnection queue projects in NJ\n"
            "- Any stranded capacity or load loss reports from NJ utilities\n\n"
            "Return ONLY this JSON:\n"
            '{{\n'
            '  "findings": [\n'
            '    {{\n'
            '      "facility_name": "...",\n'
            '      "company": "...",\n'
            '      "address": "if available",\n'
            '      "city": "...",\n'
            '      "county": "...",\n'
            '      "estimated_mw": number,\n'
            '      "utility_territory": "PSEG" | "JCP&L" | "other",\n'
            '      "type": "customer_loss" | "queue_withdrawal" | "load_reduction",\n'
            '      "date": "YYYY-MM or description",\n'
            '      "notes": "brief context",\n'
            '      "sources": ["url1"]\n'
            '    }}\n'
            '  ],\n'
            '  "summary": "one sentence"\n'
            '}}'
        ).format(query=query)

        result = call_anthropic(prompt, "part1_utility", label, cache_key)
        if result and "findings" in result:
            for f in result["findings"]:
                f["state"] = "NJ"
                f["source"] = "utility_scan"
            all_findings.extend(result["findings"])

    print("  Utility scan: {} findings across {} queries".format(
        len(all_findings), len(queries)))
    return all_findings


def run_part1(dry_run=False):
    """Part 1: NJ Industrial Closure Scan."""
    print()
    print("=" * 70)
    print("PART 1: NJ Industrial Closure Scan")
    print("=" * 70)

    if dry_run:
        print("  [DRY RUN] Would scrape NJ WARN Act, run 4 news queries, 4 utility queries")
        return []

    all_sites = []

    # 1a. WARN Act
    warn_results = scrape_nj_warn()
    for r in warn_results:
        city = r.get("city", "")
        if not city:
            continue
        coords = geocode_cached(city, "NJ")
        if not coords:
            print("    SKIP (no geocode): {}".format(r["company"]))
            continue

        est_mw = estimate_mw(r["company"], r["employees"])
        sub = find_nearest_substation(coords[0], coords[1])

        all_sites.append({
            "name": r["company"],
            "source": "warn_act",
            "source_type": "WARN Act Filing",
            "sub_type": classify_sub_type(r["company"]),
            "city": city,
            "county": "",
            "state": "NJ",
            "address": r["location"],
            "lat": coords[0],
            "lon": coords[1],
            "estimated_mw": est_mw,
            "employee_count": r["employees"],
            "closure_date": r["notice_date"],
            "closure_status": "closing" if r["is_closure"] else "announced",
            "utility_territory": "",
            "nearest_sub_name": sub["name"] if sub else "",
            "nearest_sub_miles": sub["distance_miles"] if sub else 999,
            "nearest_sub_kv": sub["max_volt"] if sub else 0,
            "sources": [NJ_WARN_URL],
            "notes": "WARN Act filing. {} employees affected.".format(r["employees"]),
            "priority": "HIGH" if est_mw >= 20 else "MEDIUM" if est_mw >= 10 else "LOW",
        })
        print("    {} — {}, ~{}MW, {} emp".format(
            r["company"][:40], city, est_mw, r["employees"]))

    # 1b. News scan
    news_closures = run_nj_news_scan()
    for c in news_closures:
        city = c.get("city", c.get("address", ""))
        if not city:
            continue

        address = c.get("address", city + ", NJ")
        coords = geocode_cached(address, "NJ")
        if not coords and city:
            coords = geocode_cached(city, "NJ")
        if not coords:
            print("    SKIP (no geocode): {}".format(c.get("facility_name", "?")))
            continue

        try:
            est_mw = int(float(c.get("estimated_mw") or 0)) or estimate_mw(
                c.get("facility_name", "") + " " + c.get("sub_type", ""),
                int(float(c.get("employee_count") or 0)))
        except (ValueError, TypeError):
            est_mw = estimate_mw(
                c.get("facility_name", "") + " " + c.get("sub_type", ""),
                int(float(c.get("employee_count") or 0)))
        sub = find_nearest_substation(coords[0], coords[1])

        all_sites.append({
            "name": c.get("facility_name", c.get("company", "Unknown")),
            "source": "news_scan",
            "source_type": "News/Research",
            "sub_type": c.get("sub_type", classify_sub_type(c.get("facility_name", ""))),
            "city": c.get("city", ""),
            "county": c.get("county", ""),
            "state": "NJ",
            "address": address,
            "lat": coords[0],
            "lon": coords[1],
            "estimated_mw": est_mw,
            "employee_count": c.get("employee_count") or 0,
            "closure_date": c.get("closure_date", ""),
            "closure_status": c.get("status", "announced"),
            "utility_territory": c.get("utility_territory", ""),
            "nearest_sub_name": sub["name"] if sub else "",
            "nearest_sub_miles": sub["distance_miles"] if sub else 999,
            "nearest_sub_kv": sub["max_volt"] if sub else 0,
            "sources": c.get("sources", []),
            "notes": c.get("notes", ""),
            "priority": "HIGH" if (est_mw or 0) >= 20 else "MEDIUM" if (est_mw or 0) >= 10 else "LOW",
        })

    # 1c. Utility load loss
    utility_findings = run_nj_utility_scan()
    for f in utility_findings:
        city = f.get("city", "")
        address = f.get("address", "")
        if not city and not address:
            continue

        coords = None
        if address:
            coords = geocode_cached(address, "NJ")
        if not coords and city:
            coords = geocode_cached(city, "NJ")
        if not coords:
            print("    SKIP (no geocode): {}".format(f.get("facility_name", "?")))
            continue

        try:
            est_mw = int(float(f.get("estimated_mw") or 0)) or 10
        except (ValueError, TypeError):
            est_mw = 10
        sub = find_nearest_substation(coords[0], coords[1])

        all_sites.append({
            "name": f.get("facility_name", f.get("company", "Unknown")),
            "source": "utility_scan",
            "source_type": "Utility Load Loss",
            "sub_type": classify_sub_type(f.get("facility_name", "")),
            "city": city,
            "county": f.get("county", ""),
            "state": "NJ",
            "address": address or (city + ", NJ"),
            "lat": coords[0],
            "lon": coords[1],
            "estimated_mw": est_mw,
            "employee_count": 0,
            "closure_date": f.get("date", ""),
            "closure_status": f.get("type", "load_reduction"),
            "utility_territory": f.get("utility_territory", ""),
            "nearest_sub_name": sub["name"] if sub else "",
            "nearest_sub_miles": sub["distance_miles"] if sub else 999,
            "nearest_sub_kv": sub["max_volt"] if sub else 0,
            "sources": f.get("sources", []),
            "notes": f.get("notes", ""),
            "priority": "HIGH" if (est_mw or 0) >= 20 else "MEDIUM" if (est_mw or 0) >= 10 else "LOW",
        })

    # Deduplicate by name similarity
    seen = set()
    deduped = []
    for s in all_sites:
        key = s["name"].lower().strip()[:30]
        if key not in seen:
            seen.add(key)
            deduped.append(s)

    all_sites = deduped

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(CLOSURES_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 1,
                "description": "NJ industrial closures (WARN Act + news + utility)",
                "total_sites": len(all_sites),
                "high_priority": sum(1 for s in all_sites if s["priority"] == "HIGH"),
                "api_spend": get_total_spend(),
            },
            "sites": all_sites,
        }, f, indent=2)

    print()
    print("  Part 1 Summary:")
    print("    Total NJ closure sites: {}".format(len(all_sites)))
    print("    HIGH priority (20MW+):  {}".format(sum(1 for s in all_sites if s["priority"] == "HIGH")))
    print("    MEDIUM priority (10-20MW): {}".format(sum(1 for s in all_sites if s["priority"] == "MEDIUM")))
    print("    API spend so far: ${:.2f}".format(get_total_spend()))
    print("    Output: {}".format(CLOSURES_FILE))

    return all_sites


# ══════════════════════════════════════════════════════════════════════════════
# PART 2: NJ Retired/Retiring Power Plant Sites
# ══════════════════════════════════════════════════════════════════════════════


def run_part2(dry_run=False):
    """Part 2: NJ Retired/Retiring Power Plants from EIA data."""
    print()
    print("=" * 70)
    print("PART 2: NJ Retired/Retiring Power Plant Sites")
    print("=" * 70)

    if not os.path.exists(PLANTS_FILE):
        print("  ERROR: power-plants.geojson not found")
        return []

    with open(PLANTS_FILE) as f:
        geo = json.load(f)

    # Filter NJ retired/retiring plants
    nj_plants = []
    for feat in geo["features"]:
        p = feat["properties"]
        if p.get("state") != "NJ":
            continue
        status = p.get("status", "")
        if status not in ("retired", "retiring"):
            continue
        coords = feat["geometry"]["coordinates"]
        nj_plants.append({
            "name": p.get("plant_name", "Unknown"),
            "state": "NJ",
            "lat": coords[1],
            "lon": coords[0],
            "capacity_mw": p.get("total_capacity_mw", 0),
            "fuel_type": p.get("fuel_type", ""),
            "status": status,
            "retirement_date": p.get("planned_retirement_date", ""),
            "owner": p.get("owner_name", ""),
        })

    print("  NJ retired/retiring plants in EIA data: {}".format(len(nj_plants)))

    if dry_run:
        for plant in nj_plants:
            print("    {} — {} MW, {}, {}".format(
                plant["name"], plant["capacity_mw"], plant["fuel_type"], plant["status"]))
        return nj_plants

    # Research each plant's current status
    enriched = []
    for plant in nj_plants:
        cache_key = "plant|{}|{}".format(plant["name"], plant["capacity_mw"])

        prompt = (
            "You are a commercial real estate research analyst specializing in power plant "
            "site redevelopment. Research this retired/retiring NJ power plant:\n\n"
            "Plant: {name}\n"
            "Capacity: {mw} MW\n"
            "Fuel Type: {fuel}\n"
            "Status: {status}\n"
            "Owner: {owner}\n"
            "Coordinates: {lat}, {lon}\n\n"
            "Research and return ONLY this JSON:\n"
            '{{\n'
            '  "current_site_status": "remediation" | "demolition" | "redevelopment" | "idle" | "active",\n'
            '  "transmission_voltage_kv": number or null,\n'
            '  "acquirability": "high" | "medium" | "low" | "institutional",\n'
            '  "acquirability_notes": "why this rating",\n'
            '  "owner_type": "utility" | "private_equity" | "municipality" | "developer" | "other",\n'
            '  "current_owner": "name",\n'
            '  "redevelopment_plans": "description or null",\n'
            '  "environmental_status": "clean" | "remediation_active" | "remediation_complete" | "unknown",\n'
            '  "partnership_path": "description of how Brian could access this site",\n'
            '  "data_center_suitability": "high" | "medium" | "low",\n'
            '  "notes": "key context",\n'
            '  "sources": ["url1"]\n'
            '}}'
        ).format(
            name=plant["name"], mw=plant["capacity_mw"], fuel=plant["fuel_type"],
            status=plant["status"], owner=plant["owner"],
            lat=plant["lat"], lon=plant["lon"])

        result = call_anthropic(prompt, "part2_plants", plant["name"][:30], cache_key)

        sub = find_nearest_substation(plant["lat"], plant["lon"])

        entry = dict(plant)
        entry["source"] = "eia_plants"
        entry["source_type"] = "EIA Retired Plant"
        entry["sub_type"] = "Retired Power Plant"
        entry["address"] = ""
        entry["city"] = ""
        entry["county"] = ""
        entry["nearest_sub_name"] = sub["name"] if sub else ""
        entry["nearest_sub_miles"] = sub["distance_miles"] if sub else 999
        entry["nearest_sub_kv"] = sub["max_volt"] if sub else 0
        entry["estimated_mw"] = plant["capacity_mw"]
        entry["employee_count"] = 0
        entry["closure_date"] = plant["retirement_date"]
        entry["closure_status"] = plant["status"]
        entry["utility_territory"] = ""
        entry["sources"] = []
        entry["notes"] = ""

        if result and not result.get("parse_error"):
            entry["current_site_status"] = result.get("current_site_status", "unknown")
            entry["transmission_kv"] = result.get("transmission_voltage_kv")
            entry["acquirability"] = result.get("acquirability", "unknown")
            entry["acquirability_notes"] = result.get("acquirability_notes", "")
            entry["owner_type"] = result.get("owner_type", "unknown")
            entry["current_owner"] = result.get("current_owner", plant["owner"])
            entry["redevelopment_plans"] = result.get("redevelopment_plans")
            entry["environmental_status"] = result.get("environmental_status", "unknown")
            entry["partnership_path"] = result.get("partnership_path", "")
            entry["dc_suitability"] = result.get("data_center_suitability", "unknown")
            entry["sources"] = result.get("sources", [])
            entry["notes"] = result.get("notes", "")

            # Flag PSEG institutional assets per spec
            is_pseg = "pseg" in (plant["owner"] or "").lower()
            is_large = plant["capacity_mw"] >= 150
            if is_pseg and is_large:
                entry["acquirability"] = "institutional"
                entry["acquirability_notes"] = (
                    "PSEG-owned institutional asset ({}MW). Not directly optionable. "
                    "Brian's BPU contacts could facilitate introduction/partnership. "
                    "{}".format(plant["capacity_mw"], entry.get("acquirability_notes", "")))

        # Priority based on capacity and acquirability
        acq = entry.get("acquirability", "unknown")
        if acq == "institutional":
            entry["priority"] = "INSTITUTIONAL"
        elif plant["capacity_mw"] >= 100:
            entry["priority"] = "HIGH"
        elif plant["capacity_mw"] >= 50:
            entry["priority"] = "MEDIUM"
        else:
            entry["priority"] = "LOW"

        enriched.append(entry)
        print("    {} — {}MW, {}, acq={}".format(
            plant["name"][:35], plant["capacity_mw"],
            entry.get("current_site_status", "?"), entry.get("acquirability", "?")))

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(PLANTS_NJ_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 2,
                "description": "NJ retired/retiring power plants with acquirability assessment",
                "total_plants": len(enriched),
                "api_spend": get_total_spend(),
            },
            "sites": enriched,
        }, f, indent=2)

    print()
    print("  Part 2 Summary:")
    print("    NJ retired/retiring plants: {}".format(len(enriched)))
    for p in ["HIGH", "MEDIUM", "INSTITUTIONAL"]:
        n = sum(1 for s in enriched if s.get("priority") == p)
        if n:
            print("    {}: {}".format(p, n))
    print("    API spend so far: ${:.2f}".format(get_total_spend()))
    print("    Output: {}".format(PLANTS_NJ_FILE))

    return enriched


# ══════════════════════════════════════════════════════════════════════════════
# PART 3: NJ Substation Proximity + Vacant Industrial Parcels
# ══════════════════════════════════════════════════════════════════════════════


TARGET_COUNTIES = [
    "Middlesex", "Mercer", "Salem", "Cumberland", "Gloucester",
    "Burlington", "Camden", "Somerset", "Union", "Essex", "Hudson", "Passaic",
]


def run_part3(dry_run=False):
    """Part 3: NJ substation proximity and vacant industrial parcel analysis."""
    print()
    print("=" * 70)
    print("PART 3: NJ Substation Proximity + Vacant Industrial Parcels")
    print("=" * 70)

    nj_subs = load_nj_substations(138)
    print("  NJ substations (138kV+): {}".format(len(nj_subs)))

    # Filter to target counties
    target_subs = []
    for s in nj_subs:
        county = s.get("county", "")
        # Include if county matches OR if no county data (search by proximity)
        if not county or any(c.lower() in county.lower() for c in TARGET_COUNTIES):
            target_subs.append(s)

    print("  In target counties: {}".format(len(target_subs)))

    if dry_run:
        for s in target_subs[:10]:
            print("    {} — {}kV, {}, {}".format(
                s["name"][:30], s["max_volt"], s.get("city", ""), s.get("county", "")))
        if len(target_subs) > 10:
            print("    ... and {} more".format(len(target_subs) - 10))
        return []

    # Group substations by area to batch API queries (save budget)
    # Research vacant parcels near clusters of substations
    county_groups = {}
    for s in target_subs:
        county = s.get("county", "Unknown")
        if county not in county_groups:
            county_groups[county] = []
        county_groups[county].append(s)

    all_parcels = []

    for county, subs in county_groups.items():
        if not county or county == "Unknown":
            continue

        cache_key = "parcels|NJ|{}".format(county)

        # List the substations for context
        sub_list = ", ".join("{} ({}kV)".format(s["name"][:20], int(s["max_volt"]))
                           for s in subs[:5])

        prompt = (
            "You are a commercial real estate research analyst specializing in industrial "
            "property acquisition in New Jersey.\n\n"
            "Research vacant or underutilized industrial parcels in {county} County, NJ "
            "that are within 2 miles of major electrical substations.\n\n"
            "Key substations in {county} County:\n{subs}\n\n"
            "Identify parcels with these motivated-seller signals:\n"
            "- Tax delinquent properties\n"
            "- Estate/trust ownership\n"
            "- Out-of-state owners\n"
            "- Long hold periods (20+ years, same owner)\n"
            "- Properties assessed below $500K\n"
            "- Vacant or underutilized industrial land\n\n"
            "Return ONLY this JSON:\n"
            '{{\n'
            '  "parcels": [\n'
            '    {{\n'
            '      "address": "...",\n'
            '      "city": "...",\n'
            '      "county": "{county}",\n'
            '      "acres": number or null,\n'
            '      "current_use": "vacant" | "underutilized" | "industrial" | "warehouse",\n'
            '      "zoning": "industrial" | "commercial" | "mixed" | "unknown",\n'
            '      "owner_type": "private" | "estate" | "trust" | "municipal" | "corporate" | "out_of_state",\n'
            '      "owner_name": "if known",\n'
            '      "assessed_value": number or null,\n'
            '      "tax_status": "current" | "delinquent" | "unknown",\n'
            '      "years_held": number or null,\n'
            '      "motivated_seller_signals": ["list of signals"],\n'
            '      "nearest_substation": "name",\n'
            '      "distance_to_substation_miles": number,\n'
            '      "notes": "...",\n'
            '      "sources": ["url"]\n'
            '    }}\n'
            '  ]\n'
            '}}'
        ).format(county=county, subs=sub_list)

        result = call_anthropic(prompt, "part3_parcels", county + " Co. parcels", cache_key)

        if result and "parcels" in result:
            for p in result["parcels"]:
                p["state"] = "NJ"
                p["county"] = county

                # Geocode the parcel
                addr = p.get("address", "")
                city = p.get("city", "")
                coords = None
                if addr:
                    coords = geocode_cached(addr, "NJ")
                if not coords and city:
                    coords = geocode_cached(city + ", " + county + " County", "NJ")

                if coords:
                    p["lat"] = coords[0]
                    p["lon"] = coords[1]
                    sub = find_nearest_substation(coords[0], coords[1])
                    if sub:
                        p["nearest_sub_name"] = sub["name"]
                        p["nearest_sub_miles"] = sub["distance_miles"]
                        p["nearest_sub_kv"] = sub["max_volt"]
                else:
                    p["lat"] = None
                    p["lon"] = None

                p["source"] = "substation_proximity"
                p["source_type"] = "Vacant Parcel near Substation"

            all_parcels.extend(result["parcels"])
            print("    {} County: {} parcels".format(county, len(result["parcels"])))

    # Save substations data too
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(SUBSTATIONS_NJ_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 3,
                "description": "NJ substations 138kV+ in target counties",
                "total_substations": len(target_subs),
            },
            "substations": target_subs,
        }, f, indent=2)

    with open(PARCELS_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 3,
                "description": "NJ vacant industrial parcels near 138kV+ substations",
                "total_parcels": len(all_parcels),
                "counties_searched": list(county_groups.keys()),
                "api_spend": get_total_spend(),
            },
            "parcels": all_parcels,
        }, f, indent=2)

    print()
    print("  Part 3 Summary:")
    print("    NJ substations (138kV+): {}".format(len(target_subs)))
    print("    Vacant parcels found: {}".format(len(all_parcels)))
    print("    API spend so far: ${:.2f}".format(get_total_spend()))
    print("    Output: {}".format(PARCELS_FILE))

    return all_parcels


# ══════════════════════════════════════════════════════════════════════════════
# PART 4: Transportation Corridor Overlay
# ══════════════════════════════════════════════════════════════════════════════

# NJ highway interchange reference points (key interchanges)
NJ_HIGHWAYS = {
    "NJ Turnpike": [
        (40.7128, -74.1745, "Exit 14 (Newark)"),
        (40.5652, -74.2854, "Exit 11 (Woodbridge)"),
        (40.4774, -74.3023, "Exit 9 (New Brunswick)"),
        (40.2835, -74.5530, "Exit 7A (Trenton)"),
        (40.0793, -74.7278, "Exit 6 (PA Turnpike)"),
        (39.8348, -75.0652, "Exit 3 (Salem)"),
        (39.6876, -75.1847, "Exit 1 (Deepwater)"),
    ],
    "Garden State Parkway": [
        (40.9176, -74.1719, "Exit 159 (Clifton)"),
        (40.7478, -74.1892, "Exit 145 (Newark)"),
        (40.5548, -74.2848, "Exit 127 (Elizabeth)"),
        (40.3274, -74.2928, "Exit 109 (Asbury)"),
        (39.9554, -74.1980, "Exit 63 (Barnegat)"),
        (39.3648, -74.4388, "Exit 17 (Atlantic City)"),
    ],
    "I-78": [
        (40.6397, -74.2150, "Newark Airport"),
        (40.6563, -74.3489, "Springfield"),
        (40.6310, -74.7640, "Clinton"),
    ],
    "I-80": [
        (40.8568, -74.2264, "Parsippany"),
        (40.9028, -74.3398, "Rockaway"),
        (40.9176, -74.5668, "Netcong"),
        (40.9734, -74.9327, "Columbia"),
    ],
    "I-95": [
        (40.2171, -74.7429, "Trenton"),
        (40.7128, -74.0060, "George Washington Bridge"),
    ],
    "I-287": [
        (40.5879, -74.6129, "Bridgewater"),
        (40.6989, -74.3991, "Morristown"),
        (40.8420, -74.2890, "Parsippany"),
        (40.9840, -74.1335, "Oakland"),
    ],
    "I-195": [
        (40.2203, -74.7547, "Trenton"),
        (40.2185, -74.5905, "Hamilton"),
        (40.1940, -74.2637, "Freehold"),
    ],
}

# NJ airports
NJ_AIRPORTS = [
    (40.6895, -74.1745, "Newark Liberty (EWR)"),
    (39.4576, -74.5773, "Atlantic City (ACY)"),
    (40.0799, -74.5946, "Trenton-Mercer (TTN)"),
    (40.9712, -74.2924, "Morristown Airport (MMU)"),
    (40.5160, -74.3507, "Linden Airport"),
]


def calculate_transport_distances(lat, lon):
    """Calculate distances to all NJ transportation infrastructure."""
    result = {}

    # Nearest point on each highway
    for highway, points in NJ_HIGHWAYS.items():
        best_dist = float("inf")
        best_exit = ""
        for plat, plon, name in points:
            d = haversine_miles(lat, lon, plat, plon)
            if d < best_dist:
                best_dist = d
                best_exit = name
        result[highway] = {"distance_miles": round(best_dist, 1), "nearest": best_exit}

    # Nearest highway overall
    all_highway_dists = []
    for highway, info in result.items():
        all_highway_dists.append((info["distance_miles"], highway, info["nearest"]))
    all_highway_dists.sort()

    if all_highway_dists:
        result["nearest_highway"] = {
            "name": all_highway_dists[0][1],
            "distance_miles": all_highway_dists[0][0],
            "interchange": all_highway_dists[0][2],
        }
    result["transportation_advantaged"] = (
        all_highway_dists[0][0] <= 1.0 if all_highway_dists else False)

    # Nearest airport
    best_airport_dist = float("inf")
    best_airport = ""
    for alat, alon, name in NJ_AIRPORTS:
        d = haversine_miles(lat, lon, alat, alon)
        if d < best_airport_dist:
            best_airport_dist = d
            best_airport = name
    result["nearest_airport"] = {"name": best_airport, "distance_miles": round(best_airport_dist, 1)}

    # Fiber route estimate (follows highways — use nearest highway as proxy)
    if all_highway_dists:
        result["estimated_fiber_distance_miles"] = all_highway_dists[0][0]
    else:
        result["estimated_fiber_distance_miles"] = 999

    return result


def run_part4(dry_run=False):
    """Part 4: Transportation Corridor Overlay."""
    print()
    print("=" * 70)
    print("PART 4: Transportation Corridor Overlay")
    print("=" * 70)

    # Load all sites from Parts 1-3
    all_sites = []
    for filepath, key in [
        (CLOSURES_FILE, "sites"),
        (PLANTS_NJ_FILE, "sites"),
        (PARCELS_FILE, "parcels"),
    ]:
        if os.path.exists(filepath):
            with open(filepath) as f:
                data = json.load(f)
            items = data.get(key, [])
            all_sites.extend(items)
            print("  Loaded {} sites from {}".format(len(items), os.path.basename(filepath)))

    if not all_sites:
        print("  ERROR: No sites loaded. Run Parts 1-3 first.")
        return []

    if dry_run:
        print("  [DRY RUN] Would calculate transport distances for {} sites".format(len(all_sites)))
        return all_sites

    # Calculate transport distances for each site
    for site in all_sites:
        lat = site.get("lat")
        lon = site.get("lon")
        if lat is None or lon is None:
            site["transport"] = {}
            continue
        site["transport"] = calculate_transport_distances(lat, lon)

    # Save combined file
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(ALL_SITES_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 4,
                "description": "All NJ sites with transportation overlay",
                "total_sites": len(all_sites),
                "transport_advantaged": sum(
                    1 for s in all_sites
                    if s.get("transport", {}).get("transportation_advantaged")),
            },
            "sites": all_sites,
        }, f, indent=2)

    ta_count = sum(1 for s in all_sites if s.get("transport", {}).get("transportation_advantaged"))
    print()
    print("  Part 4 Summary:")
    print("    Total sites with transport data: {}".format(len(all_sites)))
    print("    Transportation-advantaged (<1mi highway): {}".format(ta_count))
    print("    Output: {}".format(ALL_SITES_FILE))

    return all_sites


# ══════════════════════════════════════════════════════════════════════════════
# PART 5: Scoring and Ranking
# ══════════════════════════════════════════════════════════════════════════════


def score_site(site):
    """Score a site on the deal-book-weighted 5-dimension model.

    Weights tuned for Brian/Ralph acquirability:
      Acquirability     30%  — Can Brian and Ralph actually get this?
      Power Availability 20% — Is there confirmed or estimated stranded capacity?
      Transmission       15% — Distance to 138kV+ substation
      Transportation     15% — Highway, fiber, airport proximity
      Site Readiness     20% — Condition, recency, environmental
    """

    est_mw = site.get("estimated_mw", site.get("capacity_mw", 0)) or 0
    closure_status = site.get("closure_status", "")
    source = site.get("source", "")

    # 1. Acquirability (30%) — the most important dimension for a deal book
    acq = site.get("acquirability", "")
    owner_type = site.get("owner_type", "")
    tax_status = site.get("tax_status", "")
    motivated = site.get("motivated_seller_signals", [])
    owner = (site.get("current_owner") or site.get("owner") or "").lower()

    if acq == "high" or tax_status == "delinquent":
        acq_score = 95
    elif owner_type in ("estate", "trust", "out_of_state"):
        acq_score = 90
    elif source in ("warn_act", "news_scan"):
        # WARN/news closures — company is leaving, site will need disposition
        acq_score = 85
    elif source == "utility_scan":
        acq_score = 75
    elif acq == "medium" and owner_type == "private_equity":
        acq_score = 70
    elif owner_type == "private" or "llc" in owner:
        # Private LLC ownership — direct negotiation possible
        acq_score = 75
    elif owner_type == "municipal":
        acq_score = 55
    elif acq == "institutional" or owner_type == "utility":
        # PSEG/utility-owned — hard for Brian to option directly
        acq_score = 25
    elif source == "substation_proximity":
        n_signals = len(motivated) if isinstance(motivated, list) else 0
        acq_score = min(90, 40 + n_signals * 12)
    elif acq == "medium":
        # Generic "medium" from enrichment — penalize unknown
        acq_score = 50
    else:
        acq_score = 40

    # Penalize if owner is a major utility (GenOn, Calpine, PSEG, NRG, ConEd)
    major_utilities = ["pseg", "genon", "calpine", "nrg", "consolidated edison",
                       "con edison", "exelon", "nextera", "duke", "dominion"]
    if any(u in owner for u in major_utilities) and source == "eia_plants":
        acq_score = min(acq_score, 40)

    # 2. Power Availability (20%)
    power_score = 0
    if est_mw >= 50:
        power_score = 95
    elif est_mw >= 30:
        power_score = 80
    elif est_mw >= 20:
        power_score = 70
    elif est_mw >= 10:
        power_score = 55
    else:
        power_score = 30

    if closure_status in ("closed", "retired"):
        power_score = min(100, power_score + 5)
    if source == "utility_scan":
        power_score = min(100, power_score + 5)

    # 3. Transmission Proximity (15%)
    sub_miles = site.get("nearest_sub_miles", 999)
    sub_kv = site.get("nearest_sub_kv", 0)

    if sub_miles <= 0.5:
        trans_score = 100
    elif sub_miles <= 1:
        trans_score = 90
    elif sub_miles <= 2:
        trans_score = 75
    elif sub_miles <= 5:
        trans_score = 60
    elif sub_miles <= 10:
        trans_score = 40
    else:
        trans_score = 20

    if sub_kv >= 345:
        trans_score = min(100, trans_score + 10)
    elif sub_kv >= 230:
        trans_score = min(100, trans_score + 5)

    # 4. Transportation Access (15%)
    transport = site.get("transport", {})
    nearest_hwy = transport.get("nearest_highway", {})
    hwy_dist = nearest_hwy.get("distance_miles", 999)
    airport_dist = transport.get("nearest_airport", {}).get("distance_miles", 999)
    is_ta = transport.get("transportation_advantaged", False)

    if is_ta:
        trans_access_score = 95
    elif hwy_dist <= 2:
        trans_access_score = 80
    elif hwy_dist <= 5:
        trans_access_score = 65
    elif hwy_dist <= 10:
        trans_access_score = 50
    else:
        trans_access_score = 30

    if airport_dist <= 10:
        trans_access_score = min(100, trans_access_score + 5)

    # 5. Site Readiness (20%) — includes recency
    current_status = site.get("current_site_status", "")
    current_use = site.get("current_use", "")
    env_status = site.get("environmental_status", "")

    if current_status == "idle" or current_use == "vacant":
        readiness_score = 85
    elif current_status == "demolition" or current_use == "underutilized":
        readiness_score = 70
    elif current_status in ("remediation",) or env_status == "remediation_active":
        readiness_score = 35
    elif current_status == "redevelopment":
        readiness_score = 25  # Already claimed
    elif source in ("warn_act", "news_scan"):
        readiness_score = 75  # Recent closure — site likely still in transition
    elif source == "eia_plants":
        readiness_score = 55
    else:
        readiness_score = 50

    # Recency penalty — plants retired 5+ years ago likely already claimed or problematic
    retirement_date = site.get("retirement_date", site.get("closure_date", ""))
    if retirement_date:
        try:
            year = int(retirement_date[:4])
            years_ago = 2026 - year
            if years_ago >= 10:
                readiness_score = max(10, readiness_score - 30)
            elif years_ago >= 7:
                readiness_score = max(15, readiness_score - 20)
            elif years_ago >= 5:
                readiness_score = max(20, readiness_score - 10)
            elif years_ago <= 2:
                readiness_score = min(100, readiness_score + 10)  # Recent — bonus
        except (ValueError, IndexError):
            pass

    # Composite — deal-book weights
    composite = (
        acq_score * 0.30 +
        power_score * 0.20 +
        trans_score * 0.15 +
        trans_access_score * 0.15 +
        readiness_score * 0.20
    )

    return {
        "composite_score": round(composite, 1),
        "acquirability_score": round(acq_score, 1),
        "power_availability": round(power_score, 1),
        "transmission_proximity": round(trans_score, 1),
        "transportation_access": round(trans_access_score, 1),
        "site_readiness": round(readiness_score, 1),
    }


def run_part5(dry_run=False):
    """Part 5: Score and rank all NJ sites, select top 10."""
    print()
    print("=" * 70)
    print("PART 5: Scoring and Ranking")
    print("=" * 70)

    if not os.path.exists(ALL_SITES_FILE):
        print("  ERROR: {} not found. Run Part 4 first.".format(ALL_SITES_FILE))
        return []

    with open(ALL_SITES_FILE) as f:
        data = json.load(f)
    all_sites = data.get("sites", [])

    print("  Scoring {} sites...".format(len(all_sites)))

    if dry_run:
        print("  [DRY RUN] Would score {} sites and select top 10".format(len(all_sites)))
        return []

    # Score each site
    for site in all_sites:
        scores = score_site(site)
        site["scores"] = scores

    # Sort by composite score
    all_sites.sort(key=lambda s: -s.get("scores", {}).get("composite_score", 0))

    # Select top 10
    top10 = all_sites[:10]

    # Save scored results
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(SCORED_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 5,
                "description": "NJ sites scored and ranked",
                "total_scored": len(all_sites),
                "top_10_selected": len(top10),
                "scoring_model": {
                    "power_availability": "30%",
                    "transmission_proximity": "20%",
                    "acquirability": "20%",
                    "transportation_access": "15%",
                    "site_readiness": "15%",
                },
            },
            "all_ranked": all_sites,
            "top_10": top10,
        }, f, indent=2)

    print()
    print("  Part 5 Results — Top 10 NJ Sites:")
    print("  " + "-" * 68)
    print("  {:>3} {:30s} {:>6s} {:>6s} {:>6s} {:>5s}".format(
        "#", "Site", "Score", "Power", "Trans", "Acq"))
    print("  " + "-" * 68)
    for i, site in enumerate(top10):
        s = site["scores"]
        print("  {:>3} {:30s} {:6.1f} {:6.1f} {:6.1f} {:5.1f}".format(
            i + 1,
            (site.get("name", "?"))[:30],
            s["composite_score"],
            s["power_availability"],
            s["transmission_proximity"],
            s["acquirability_score"],
        ))
    print("  " + "-" * 68)
    print("  Output: {}".format(SCORED_FILE))

    return top10


# ══════════════════════════════════════════════════════════════════════════════
# PART 6: Generate Site Briefs
# ══════════════════════════════════════════════════════════════════════════════


def generate_site_brief(site, rank):
    """Generate a detailed one-page brief for a site using API research."""
    name = site.get("name", "Unknown")
    cache_key = "brief|{}|{}".format(name[:30], rank)

    scores = site.get("scores", {})
    transport = site.get("transport", {})
    nearest_hwy = transport.get("nearest_highway", {})

    prompt = (
        "You are a data center real estate advisor preparing a deal book for a meeting "
        "with a multifamily developer (Ralph) who has capital, NJ Turnpike/DOT infrastructure "
        "background, and trust with Brian (10+ years institutional DC real estate at JLL, Goodman). "
        "Ralph understands data center opportunity but has no DC development experience.\n\n"
        "Generate a comprehensive site brief for this NJ data center opportunity:\n\n"
        "Site: {name}\n"
        "Location: {address}\n"
        "Type: {sub_type}\n"
        "Source: {source_type}\n"
        "Estimated Power: {mw} MW\n"
        "Employees: {emp}\n"
        "Closure Date: {closure}\n"
        "Status: {status}\n"
        "Utility Territory: {utility}\n"
        "Nearest Substation: {sub_name} ({sub_kv}kV, {sub_mi} mi)\n"
        "Nearest Highway: {hwy_name} ({hwy_mi} mi)\n"
        "Nearest Airport: {airport}\n"
        "Composite Score: {score}/100\n"
        "Owner: {owner}\n"
        "Acquirability: {acq}\n"
        "County: {county}\n\n"
        "Return ONLY this JSON:\n"
        '{{\n'
        '  "opportunity_summary": "2-3 sentences: what is this site, why power is available, why now",\n'
        '  "power_story": {{\n'
        '    "estimated_capacity_mw": number,\n'
        '    "source_of_capacity": "industrial closure | retired plant | withdrawn queue | substation proximity",\n'
        '    "utility_territory": "PSEG | JCP&L | other",\n'
        '    "transmission": "nearest substation details",\n'
        '    "confidence_level": "High | Medium | Speculative"\n'
        '  }},\n'
        '  "site_details": {{\n'
        '    "parcel_size_acres": number or null,\n'
        '    "current_use": "...",\n'
        '    "zoning": "...",\n'
        '    "owner_type": "private | estate | institutional | municipal",\n'
        '    "estimated_value": "range",\n'
        '    "environmental_flags": "description or none"\n'
        '  }},\n'
        '  "acquirability": {{\n'
        '    "owner_motivation": "description",\n'
        '    "estimated_acquisition_cost": "range",\n'
        '    "recommended_approach": "direct purchase | option | partnership | broker",\n'
        '    "notes": "..."\n'
        '  }},\n'
        '  "developer_value": {{\n'
        '    "estimated_developer_price": "range",\n'
        '    "spread_estimate": "range",\n'
        '    "comparable_transactions": "if any"\n'
        '  }},\n'
        '  "next_steps": [\n'
        '    "specific action 1",\n'
        '    "specific action 2",\n'
        '    "specific action 3"\n'
        '  ],\n'
        '  "ralph_should_see": true | false,\n'
        '  "bpu_contact_relevant": "description or null"\n'
        '}}'
    ).format(
        name=name,
        address=site.get("address", site.get("city", "") + ", NJ"),
        sub_type=site.get("sub_type", ""),
        source_type=site.get("source_type", ""),
        mw=site.get("estimated_mw", site.get("capacity_mw", "?")),
        emp=site.get("employee_count", "?"),
        closure=site.get("closure_date", "?"),
        status=site.get("closure_status", site.get("current_site_status", "?")),
        utility=site.get("utility_territory", "?"),
        sub_name=site.get("nearest_sub_name", "?"),
        sub_kv=site.get("nearest_sub_kv", "?"),
        sub_mi=site.get("nearest_sub_miles", "?"),
        hwy_name=nearest_hwy.get("name", "?"),
        hwy_mi=nearest_hwy.get("distance_miles", "?"),
        airport=transport.get("nearest_airport", {}).get("name", "?"),
        score=scores.get("composite_score", "?"),
        owner=site.get("current_owner", site.get("owner", "?")),
        acq=site.get("acquirability", site.get("acquirability_notes", "?")),
        county=site.get("county", "?"),
    )

    result = call_anthropic(prompt, "part6_briefs", "Brief: " + name[:30], cache_key)
    return result


def run_part6(dry_run=False):
    """Part 6: Generate site briefs for top 10."""
    print()
    print("=" * 70)
    print("PART 6: Generate NJ Site Briefs")
    print("=" * 70)

    if not os.path.exists(SCORED_FILE):
        print("  ERROR: {} not found. Run Part 5 first.".format(SCORED_FILE))
        return []

    with open(SCORED_FILE) as f:
        data = json.load(f)
    top10 = data.get("top_10", [])

    print("  Generating briefs for {} sites...".format(len(top10)))

    if dry_run:
        for i, site in enumerate(top10):
            print("    {}. {} — score {:.1f}".format(
                i + 1, site.get("name", "?"),
                site.get("scores", {}).get("composite_score", 0)))
        return top10

    briefs = []
    for i, site in enumerate(top10):
        print("\n  [{}/{}] {}".format(i + 1, len(top10), site.get("name", "?")))
        brief = generate_site_brief(site, i + 1)
        site["brief"] = brief or {}
        briefs.append(site)

    # Save
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(BRIEFS_FILE, "w") as f:
        json.dump({
            "metadata": {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "part": 6,
                "description": "NJ deal book site briefs",
                "total_briefs": len(briefs),
                "api_spend": get_total_spend(),
            },
            "briefs": briefs,
        }, f, indent=2)

    print()
    print("  Part 6 Summary:")
    print("    Briefs generated: {}".format(len(briefs)))
    print("    API spend so far: ${:.2f}".format(get_total_spend()))
    print("    Output: {}".format(BRIEFS_FILE))

    return briefs


# ══════════════════════════════════════════════════════════════════════════════
# PART 7: Compile Deal Book (Markdown + PDF)
# ══════════════════════════════════════════════════════════════════════════════


def generate_markdown(briefs):
    """Generate the full deal book as markdown."""
    now = datetime.now().strftime("%B %d, %Y")

    md = []
    md.append("# NJ Data Center Site Opportunities — Confidential")
    md.append("")
    md.append("**Prepared by:** SPS — Strategic Property Solutions")
    md.append("**Date:** {}".format(now))
    md.append("**Classification:** Confidential — For Authorized Recipients Only")
    md.append("")
    md.append("---")
    md.append("")

    # Executive Summary
    md.append("## Executive Summary")
    md.append("")
    md.append("This deal book presents {} high-potential data center development sites "
              "in New Jersey, identified through a systematic analysis of stranded power "
              "capacity, retired generation facilities, and vacant industrial parcels near "
              "high-voltage transmission infrastructure.".format(len(briefs)))
    md.append("")
    md.append("**The Thesis:** When large industrial consumers close — manufacturing plants, "
              "refineries, distribution centers — they leave behind utility infrastructure sized "
              "for 10-100+ MW loads. This infrastructure can take 3-5 years and $50M+ to build "
              "new. By identifying these stranded capacity sites before they're widely marketed, "
              "we can acquire sites at industrial land prices and deliver them to data center "
              "developers at a significant premium.")
    md.append("")
    md.append("**Why New Jersey:**")
    md.append("- Top-tier fiber connectivity (97% broadband coverage, dense metro fiber)")
    md.append("- Adjacent to NYC financial and cloud markets")
    md.append("- NJ Next AI incentive program ($250M/project)")
    md.append("- PSEG and JCP&L territory with available transmission capacity")
    md.append("- Competitive utility rates vs. Northern Virginia")
    md.append("- Growing hyperscaler demand pushing out of Ashburn/NoVA")
    md.append("")
    md.append("**How These Sites Were Selected:**")
    md.append("1. NJ WARN Act filings (plant closures, 100+ employees, 36 months)")
    md.append("2. Industrial closure news scan (manufacturing, chemical, distribution)")
    md.append("3. PSEG/JCP&L load loss and PJM queue withdrawal research")
    md.append("4. EIA retired/retiring power plant database")
    md.append("5. Vacant industrial parcels within 2 miles of 138kV+ substations")
    md.append("6. Scored on: Power Availability (30%), Transmission (20%), "
              "Acquirability (20%), Transportation (15%), Site Readiness (15%)")
    md.append("")
    md.append("---")
    md.append("")

    # Market Context
    md.append("## NJ Data Center Market Context")
    md.append("")
    md.append("### Demand Drivers")
    md.append("- Northern Virginia (the world's largest data center market) is running out of power")
    md.append("- Hyperscalers (AWS, Azure, Google, Meta) are actively expanding into NJ")
    md.append("- Financial services firms require NJ proximity for low-latency NYC connections")
    md.append("- AI training and inference workloads driving unprecedented MW demand")
    md.append("")
    md.append("### NJ Incentive Environment")
    md.append("- **NJ Next AI Program:** Up to $250M per qualifying project")
    md.append("- **NJEDA Tax Incentives:** Corporate business tax credits for qualifying investments")
    md.append("- **Opportunity Zones:** Multiple qualifying census tracts near identified sites")
    md.append("")
    md.append("### Utility Rate Environment")
    md.append("- PSEG average commercial rate: competitive with PJM wholesale + delivery")
    md.append("- JCP&L (FirstEnergy): competitive for large industrial loads")
    md.append("- PJM wholesale market access for large consumers (>1MW)")
    md.append("")
    md.append("---")
    md.append("")

    # Site Briefs
    md.append("## Site Opportunities")
    md.append("")

    for i, site in enumerate(briefs):
        brief = site.get("brief", {})
        scores = site.get("scores", {})
        transport = site.get("transport", {})
        nearest_hwy = transport.get("nearest_highway", {})

        name = site.get("name", "Unknown Site")
        address = site.get("address", site.get("city", "") + ", NJ")
        county = site.get("county", "")
        city = site.get("city", "")
        location_line = city
        if county:
            location_line = "{}, {} County, NJ".format(city, county)
        elif city:
            location_line = "{}, NJ".format(city)
        else:
            location_line = "NJ"

        md.append("### {}. {} ".format(i + 1, name))
        md.append("**{}**".format(address))
        md.append("*{}*".format(location_line))
        md.append("")
        md.append("**Composite Score: {:.1f}/100**".format(scores.get("composite_score", 0)))
        md.append("")

        # The Opportunity
        opp = brief.get("opportunity_summary", "")
        if opp:
            md.append("**The Opportunity:**")
            md.append(opp)
            md.append("")

        # Power Story
        ps = brief.get("power_story", {})
        md.append("**Power Story:**")
        md.append("- Estimated available capacity: {} MW".format(
            ps.get("estimated_capacity_mw", site.get("estimated_mw",
            site.get("capacity_mw", "?")))))
        md.append("- Source of capacity: {}".format(
            ps.get("source_of_capacity", site.get("source_type", "?"))))
        md.append("- Utility territory: {}".format(
            ps.get("utility_territory", site.get("utility_territory", "?"))))
        md.append("- Transmission: {}".format(
            ps.get("transmission", "{} ({}kV, {} mi)".format(
                site.get("nearest_sub_name", "?"),
                site.get("nearest_sub_kv", "?"),
                site.get("nearest_sub_miles", "?")))))
        md.append("- Confidence level: {}".format(ps.get("confidence_level", "Medium")))
        md.append("")

        # Site Details
        sd = brief.get("site_details", {})
        md.append("**Site Details:**")
        if sd.get("parcel_size_acres"):
            md.append("- Parcel size: {} acres".format(sd["parcel_size_acres"]))
        md.append("- Current use / condition: {}".format(
            sd.get("current_use", site.get("current_site_status",
            site.get("current_use", "?")))))
        md.append("- Zoning: {}".format(sd.get("zoning", "?")))
        md.append("- Owner type: {}".format(
            sd.get("owner_type", site.get("owner_type", "?"))))
        if sd.get("estimated_value"):
            md.append("- Estimated current value: {}".format(sd["estimated_value"]))
        if sd.get("environmental_flags") and sd["environmental_flags"] != "none":
            md.append("- Environmental flags: {}".format(sd["environmental_flags"]))
        md.append("")

        # Transportation
        md.append("**Transportation:**")
        md.append("- Nearest highway interchange: {} ({} mi)".format(
            nearest_hwy.get("name", "?"),
            nearest_hwy.get("distance_miles", "?")))
        md.append("- Nearest fiber route (estimated): {:.1f} mi".format(
            transport.get("estimated_fiber_distance_miles", 999)))
        md.append("- Nearest airport: {} ({} mi)".format(
            transport.get("nearest_airport", {}).get("name", "?"),
            transport.get("nearest_airport", {}).get("distance_miles", "?")))
        if transport.get("transportation_advantaged"):
            md.append("- **Transportation-advantaged** (<1 mi from major interchange)")
        md.append("")

        # Acquirability
        aq = brief.get("acquirability", {})
        md.append("**Acquirability:**")
        if aq.get("owner_motivation"):
            md.append("- Owner / motivation: {}".format(aq["owner_motivation"]))
        if aq.get("estimated_acquisition_cost"):
            md.append("- Estimated acquisition cost: {}".format(aq["estimated_acquisition_cost"]))
        if aq.get("recommended_approach"):
            md.append("- Recommended approach: {}".format(aq["recommended_approach"]))
        if aq.get("notes"):
            md.append("- Notes: {}".format(aq["notes"]))
        md.append("")

        # Developer Value
        dv = brief.get("developer_value", {})
        md.append("**Data Center Developer Value:**")
        if dv.get("estimated_developer_price"):
            md.append("- Estimated developer price: {}".format(dv["estimated_developer_price"]))
        if dv.get("spread_estimate"):
            md.append("- Estimated spread: {}".format(dv["spread_estimate"]))
        if dv.get("comparable_transactions"):
            md.append("- Comparable transactions: {}".format(dv["comparable_transactions"]))
        md.append("")

        # Next Steps
        steps = brief.get("next_steps", [])
        if steps:
            md.append("**Next Steps:**")
            for step in steps:
                md.append("- {}".format(step))
            md.append("")

        if brief.get("ralph_should_see"):
            md.append("*Ralph should see this site.*")
            md.append("")
        if brief.get("bpu_contact_relevant"):
            md.append("*BPU contact: {}*".format(brief["bpu_contact_relevant"]))
            md.append("")

        md.append("---")
        md.append("")

    # Appendix
    md.append("## Appendix")
    md.append("")
    md.append("### Methodology")
    md.append("")
    md.append("Sites were identified through three data pipelines:")
    md.append("1. **WARN Act Analysis:** NJ Department of Labor WARN filings filtered for "
              "plant closures with 100+ employees in heavy industrial sectors (36-month window)")
    md.append("2. **News & Utility Research:** Anthropic Claude API with web search to identify "
              "industrial closures, PSEG/JCP&L load loss, and PJM queue withdrawals")
    md.append("3. **Infrastructure Analysis:** EIA-860 retired plant database and HIFLD substation "
              "data filtered for NJ with 138kV+ transmission")
    md.append("")
    md.append("### Scoring Model")
    md.append("")
    md.append("| Dimension | Weight | Description |")
    md.append("|-----------|--------|-------------|")
    md.append("| Power Availability | 30% | Confirmed or estimated stranded capacity (MW) |")
    md.append("| Transmission Proximity | 20% | Distance to 138kV+ substation |")
    md.append("| Acquirability | 20% | Ownership type, motivated seller signals |")
    md.append("| Transportation Access | 15% | Highway interchange, fiber, airport proximity |")
    md.append("| Site Readiness | 15% | Current condition, environmental status |")
    md.append("")
    md.append("### Data Sources")
    md.append("")
    md.append("- NJ Department of Labor WARN Act filings")
    md.append("- EIA-860 generator database (December 2025)")
    md.append("- HIFLD substation and transmission line data")
    md.append("- PJM interconnection queue")
    md.append("- Anthropic Claude API with web search (site research)")
    md.append("- Nominatim/OpenStreetMap (geocoding)")
    md.append("")
    md.append("### Disclaimer")
    md.append("")
    md.append("This report is prepared for informational purposes only and does not constitute "
              "an offer to sell or solicitation to buy any property. All estimates of power "
              "capacity, property values, and development costs are preliminary and require "
              "independent verification. Environmental conditions, zoning approvals, and utility "
              "interconnection availability must be confirmed through appropriate due diligence. "
              "Past transactions and market data cited may not reflect current conditions.")
    md.append("")
    md.append("---")
    md.append("")
    md.append("*Generated {} by GridSite NJ Deal Book Pipeline*".format(now))

    return "\n".join(md)


def generate_pdf(markdown_text):
    """Generate PDF from the deal book content using fpdf2."""
    try:
        from fpdf import FPDF
    except ImportError:
        print("  ERROR: fpdf2 not installed (pip install fpdf2)")
        return False

    def sanitize(text):
        """Replace unicode characters that Helvetica can't render."""
        replacements = {
            "\u2014": "--",   # em-dash
            "\u2013": "-",    # en-dash
            "\u2018": "'",    # left single quote
            "\u2019": "'",    # right single quote
            "\u201c": '"',    # left double quote
            "\u201d": '"',    # right double quote
            "\u2022": "*",    # bullet
            "\u2026": "...",  # ellipsis
            "\u00a0": " ",    # non-breaking space
            "\u00b0": " deg", # degree
        }
        for old, new in replacements.items():
            text = text.replace(old, new)
        # Strip any remaining non-latin1 chars
        return text.encode("latin-1", errors="replace").decode("latin-1")

    class DealBookPDF(FPDF):
        def header(self):
            if self.page_no() > 1:
                self.set_font("Helvetica", "I", 8)
                self.set_text_color(128, 128, 128)
                self.cell(0, 5, "NJ Data Center Site Opportunities -- Confidential", align="C")
                self.ln(8)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, "Page {}".format(self.page_no()), align="C")

    pdf = DealBookPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(20, 15, 20)

    # Cover page
    pdf.add_page()
    pdf.ln(60)
    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(0, 51, 102)
    pdf.multi_cell(0, 12, sanitize("NJ Data Center\nSite Opportunities"), align="C")
    pdf.ln(10)
    pdf.set_font("Helvetica", "", 16)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, sanitize("Confidential"), align="C")
    pdf.ln(20)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 8, sanitize("Prepared by: SPS -- Strategic Property Solutions"), align="C")
    pdf.ln(8)
    now = datetime.now().strftime("%B %d, %Y")
    pdf.cell(0, 8, sanitize("Date: {}".format(now)), align="C")

    # Parse markdown sections
    lines = markdown_text.split("\n")
    i = 0

    # Skip to after cover info (find first ## heading)
    while i < len(lines) and not lines[i].startswith("## "):
        i += 1

    pdf.add_page()

    while i < len(lines):
        line = lines[i].rstrip()

        if line.startswith("---"):
            pdf.ln(3)
            pdf.set_draw_color(200, 200, 200)
            pdf.line(20, pdf.get_y(), 190, pdf.get_y())
            pdf.ln(3)
        elif line.startswith("## "):
            # Major section header
            if pdf.get_y() > 200:
                pdf.add_page()
            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 18)
            pdf.set_text_color(0, 51, 102)
            text = sanitize(line[3:].strip())
            pdf.multi_cell(0, 9, text)
            pdf.ln(3)
        elif line.startswith("### "):
            text = sanitize(line[4:].strip())
            # Site headers get new pages
            if text and text[0].isdigit() and "." in text[:3]:
                pdf.add_page()
                pdf.set_font("Helvetica", "B", 16)
            else:
                if pdf.get_y() > 230:
                    pdf.add_page()
                pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(0, 51, 102)
            pdf.multi_cell(0, 8, text)
            pdf.ln(2)
        elif line.startswith("**") and line.endswith("**"):
            # Bold line
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(0, 0, 0)
            text = sanitize(line.strip("*").strip())
            pdf.multi_cell(0, 6, text)
            pdf.ln(1)
        elif line.startswith("**") and ":**" in line:
            # Bold label with content on same line or next
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(0, 0, 0)
            text = sanitize(line.replace("**", "").strip())
            pdf.multi_cell(0, 6, text)
            pdf.ln(1)
        elif line.startswith("- "):
            # Bullet point
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(40, 40, 40)
            text = line[2:].strip()
            # Clean markdown bold/italic
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
            text = re.sub(r'\*(.+?)\*', r'\1', text)
            text = sanitize(text)
            x = pdf.get_x()
            pdf.cell(5, 5, "")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(5, 5, "-")
            pdf.multi_cell(0, 5, " " + text)
            pdf.ln(1)
        elif line.startswith("| "):
            # Table row
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(40, 40, 40)
            cells = [c.strip() for c in line.split("|")[1:-1]]
            if all(c.replace("-", "") == "" for c in cells):
                i += 1
                continue  # Skip separator rows
            col_widths = [40, 20, 110]
            for ci, cell in enumerate(cells):
                w = col_widths[ci] if ci < len(col_widths) else 40
                is_header = i > 0 and lines[i - 1].startswith("| ")
                pdf.cell(w, 6, sanitize(cell[:int(w / 2)]), border=0)
            pdf.ln(6)
        elif line.startswith("*") and line.endswith("*") and not line.startswith("**"):
            # Italic
            pdf.set_font("Helvetica", "I", 10)
            pdf.set_text_color(80, 80, 80)
            text = sanitize(line.strip("*").strip())
            pdf.multi_cell(0, 5, text)
            pdf.ln(2)
        elif line.strip():
            # Regular paragraph
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(40, 40, 40)
            text = line.strip()
            text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
            text = re.sub(r'\*(.+?)\*', r'\1', text)
            text = sanitize(text)
            pdf.multi_cell(0, 5, text)
            pdf.ln(2)

        i += 1

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    pdf.output(DEALBOOK_PDF)
    return True


def run_part7(dry_run=False):
    """Part 7: Compile deal book as PDF and markdown."""
    print()
    print("=" * 70)
    print("PART 7: Compile NJ Deal Book")
    print("=" * 70)

    if not os.path.exists(BRIEFS_FILE):
        print("  ERROR: {} not found. Run Part 6 first.".format(BRIEFS_FILE))
        return

    with open(BRIEFS_FILE) as f:
        data = json.load(f)
    briefs = data.get("briefs", [])

    print("  Compiling deal book with {} site briefs...".format(len(briefs)))

    if dry_run:
        print("  [DRY RUN] Would generate markdown and PDF")
        return

    # Generate markdown
    md_content = generate_markdown(briefs)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(DEALBOOK_MD, "w") as f:
        f.write(md_content)
    print("  Markdown: {} ({:.1f} KB)".format(
        DEALBOOK_MD, os.path.getsize(DEALBOOK_MD) / 1024))

    # Generate PDF
    if generate_pdf(md_content):
        print("  PDF: {} ({:.1f} KB)".format(
            DEALBOOK_PDF, os.path.getsize(DEALBOOK_PDF) / 1024))
    else:
        print("  PDF generation failed")

    print()
    print("  Part 7 Summary:")
    print("    Total API spend: ${:.2f} / ${:.2f}".format(get_total_spend(), BUDGET_LIMIT))
    print("    Outputs:")
    print("      {}".format(DEALBOOK_MD))
    print("      {}".format(DEALBOOK_PDF))


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════


def main():
    parser = argparse.ArgumentParser(description="NJ Deal Book Pipeline")
    parser.add_argument("--part", type=int, help="Run a specific part (1-7)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without API calls")
    parser.add_argument("--budget", action="store_true", help="Show current API spend")
    args = parser.parse_args()

    if args.budget:
        spent = get_total_spend()
        print("NJ Deal Book API Spend: ${:.2f} / ${:.2f} ({:.1f}% of budget)".format(
            spent, BUDGET_LIMIT, spent / BUDGET_LIMIT * 100 if BUDGET_LIMIT else 0))
        return

    print("=" * 70)
    print("NJ DEAL BOOK PIPELINE")
    print("=" * 70)
    print("  Budget: ${:.2f}".format(BUDGET_LIMIT))
    print("  Output: {}".format(OUTPUT_DIR))

    init_db()

    parts = {
        1: run_part1,
        2: run_part2,
        3: run_part3,
        4: run_part4,
        5: run_part5,
        6: run_part6,
        7: run_part7,
    }

    if args.part:
        if args.part not in parts:
            print("ERROR: Invalid part number. Use 1-7.")
            sys.exit(1)
        parts[args.part](dry_run=args.dry_run)
    else:
        for part_num in sorted(parts.keys()):
            ok, remaining = check_budget()
            if not ok:
                print("\n  BUDGET EXCEEDED — stopping at Part {}".format(part_num))
                print("  Spent: ${:.2f} / ${:.2f}".format(get_total_spend(), BUDGET_LIMIT))
                break
            parts[part_num](dry_run=args.dry_run)

    print()
    print("=" * 70)
    print("COMPLETE — Total API spend: ${:.2f} / ${:.2f}".format(
        get_total_spend(), BUDGET_LIMIT))
    print("=" * 70)


if __name__ == "__main__":
    main()
